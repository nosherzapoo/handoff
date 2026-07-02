"""Build the US concert promoter market analysis workbook.

Reads cached Pollstar pages, classifies promoters (promoter_map.classify),
and writes Pollstar_US_Promoter_Analysis.xlsx with native Excel charts.
"""
import glob
import json
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.chart import AreaChart, BarChart, LineChart, Reference, Series
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from promoter_map import (AEG_BRANDS, INDIE_CANON, LN_BRANDS, NONMUSIC, classify)

YEARS = list(range(2010, 2026))            # full years 2010-2025
TIER2_CUTOFF = 93_000_000                   # ~0.1% of US period gross -> "large independent"
OUT = "Pollstar_US_Promoter_Analysis.xlsx"

# ---------- load + classify ----------
print("Loading pages ...")
ev = []
for f in sorted(glob.glob("pages/page-*.json")):
    ev += json.load(open(f))


def year_of(e):
    parts = (e.get("eventDate") or "").split("/")
    return int(parts[-1]) if len(parts) == 3 and parts[-1].isdigit() else None


def grossf(e):
    return float(e.get("grossUSD") or 0)


def ticketsf(e):
    return float(e.get("ticketsSold") or 0)


us = [e for e in ev if e.get("country") == "United States"]
rows = []
for e in us:
    y = year_of(e)
    if y is None or y < 2010 or y > 2025:
        continue
    parent, seg = classify(e.get("promoter"))
    rows.append((y, parent, seg, grossf(e), ticketsf(e)))

print(f"US 2010-2025 shows: {len(rows):,}")

# ---------- aggregate ----------
parent_gross = Counter()
parent_seg = {}
for y, p, seg, g, t in rows:
    parent_gross[p] += g
    parent_seg[p] = seg

# tier assignment
def tier_of(parent):
    seg = parent_seg[parent]
    if seg == "Major":
        return "Tier 1 (Majors)"
    if seg == "In-House":
        return "Venue / In-House"
    if seg == "Non-music":
        return "Non-music content"
    return "Tier 2 (Large indie)" if parent_gross[parent] >= TIER2_CUTOFF else "Tier 3 (Small indie)"

# yearly market overview
yr_gross = Counter(); yr_tix = Counter(); yr_shows = Counter()
yr_parents = defaultdict(set)
# parent x year gross
py_gross = defaultdict(Counter)            # py_gross[parent][year]
# segment/tier x year gross
tier_year = defaultdict(Counter)           # tier_year[tier][year]
# parent totals (shows, tix)
parent_shows = Counter(); parent_tix = Counter()
for y, p, seg, g, t in rows:
    yr_gross[y] += g; yr_tix[y] += t; yr_shows[y] += 1
    yr_parents[y].add(p)
    py_gross[p][y] += g
    tier_year[tier_of(p)][y] += g
    parent_shows[p] += 1; parent_tix[p] += t

TOTAL = sum(parent_gross.values())

# independents (Tier 2 + Tier 3) for fragmentation
indie_parents = [p for p in parent_gross if parent_seg[p] == "Independent"]

# fragmentation by year over the independent segment
frag = {}
for y in YEARS:
    shares = []
    seg_total = 0.0
    active = 0
    for p in indie_parents:
        gy = py_gross[p].get(y, 0)
        if gy > 0:
            shares.append(gy); seg_total += gy; active += 1
    shares.sort(reverse=True)
    hhi = sum((s / seg_total) ** 2 for s in shares) * 10000 if seg_total else 0
    top10 = sum(shares[:10]) / seg_total * 100 if seg_total else 0
    frag[y] = dict(active=active, hhi=hhi, top10=top10, gross=seg_total)

# ---------- workbook styling helpers ----------
wb = Workbook()
HDR = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor="1F3864")
TITLE = Font(bold=True, size=14, color="1F3864")
SUB = Font(italic=True, size=10, color="595959")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '#,##0'
MONEY_M = '#,##0.0,,"M"'
PCT = '0.0%'
NUM = '#,##0'


def style_header(ws, row, ncols, start=1):
    for c in range(start, start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autowidth(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ================= Sheet 1: README =================
ws = wb.active
ws.title = "README"
ws.sheet_view.showGridLines = False
ws["A1"] = "US Concert Promoter Market Analysis"; ws["A1"].font = TITLE
ws["A2"] = "Source: Pollstar Data Cloud boxoffice reports. Scope: United States, full years 2010-2025."
ws["A2"].font = SUB
readme = [
    ("", ""),
    ("Headline numbers", ""),
    ("US shows analyzed (2010-2025)", f"{len(rows):,}"),
    ("US total reported gross", f"${TOTAL/1e9:,.1f} billion"),
    ("Distinct promoter parents", f"{len(parent_gross):,}"),
    ("Live Nation share of US gross", f"{parent_gross['Live Nation']/TOTAL*100:.1f}%"),
    ("AEG Presents share of US gross", f"{parent_gross['AEG Presents']/TOTAL*100:.1f}%"),
    ("Independent (Tier 2 + Tier 3) share", f"{sum(parent_gross[p] for p in indie_parents)/TOTAL*100:.1f}%"),
    ("", ""),
    ("Method and definitions", ""),
    ("Metric", "Market share is measured by reported gross USD (primary). Shows and tickets shown as secondary."),
    ("Geography", "country = United States only. ~70% of all global Pollstar reports."),
    ("Window", "Full years 2010-2025. 2026 excluded (partial). Recent 1-2 years may carry Pollstar reporting lag and can understate."),
    ("Promoter roll-up", "Subsidiaries are aggregated to parent. Live Nation absorbs House of Blues, C3 Presents, Frank Productions / FPC Live. AEG Presents absorbs Goldenvoice, Concerts West, Messina Touring Group, The Bowery Presents, PromoWest."),
    ("Co-promotions", "Major-leads rule: if Live Nation or AEG appears in a co-promotion, the full show gross is credited to that major (industry lead-promoter convention). This can modestly overstate the majors."),
    ("In-House", "Pure venue self-promotion is a separate 'Venue / In-House' bucket, not counted as a promoter."),
    ("Non-music", "Sports and family/arena content owners (Feld, Cirque du Soleil, WWE/UFC/TKO, etc.) are a separate bucket so they do not distort music-promoter share."),
    ("Tier 1", "Live Nation and AEG Presents."),
    ("Tier 2", f"Independent promoters with >= ${TIER2_CUTOFF/1e6:.0f}M lifetime US gross (about 0.1% of the US market). National and super-regional players."),
    ("Tier 3", "All remaining independent promoters (regional, local, club level). This is the fragmented long tail."),
    ("Conservative choices", "Minority-stake or partial brands (Insomniac, Emporium, AC Entertainment, NS2, Marshall Arts, Madison House, 313 Presents) are left INDEPENDENT to avoid overstating the majors."),
    ("Caveat", "Touring Broadway and casino/venue operators (MSG, MGM, Caesars) appear as independents; they can be reclassified on request."),
]
r = 4
for a, b in readme:
    ws.cell(row=r, column=1, value=a).font = BOLD if b == "" and a else Font()
    ws.cell(row=r, column=2, value=b)
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
autowidth(ws, [34, 95])

# ================= Sheet 2: Market Overview by Year =================
ws = wb.create_sheet("Market Overview")
ws.sheet_view.showGridLines = False
ws["A1"] = "US Concert Market by Year (2010-2025)"; ws["A1"].font = TITLE
hdr = ["Year", "Shows", "Tickets Sold", "Gross USD", "Avg Ticket Price", "Active Promoters"]
ws.append([]); ws.append(hdr)
style_header(ws, 3, len(hdr))
for y in YEARS:
    atp = yr_gross[y] / yr_tix[y] if yr_tix[y] else 0
    ws.append([y, yr_shows[y], yr_tix[y], yr_gross[y], atp, len(yr_parents[y])])
last = ws.max_row
for rr in range(4, last + 1):
    ws.cell(rr, 2).number_format = NUM
    ws.cell(rr, 3).number_format = NUM
    ws.cell(rr, 4).number_format = MONEY
    ws.cell(rr, 5).number_format = '$#,##0.00'
    ws.cell(rr, 6).number_format = NUM
autowidth(ws, [8, 12, 16, 18, 16, 16])
# gross trend chart
ch = BarChart(); ch.title = "US Reported Gross by Year"; ch.type = "col"; ch.height = 8; ch.width = 18
data = Reference(ws, min_col=4, min_row=3, max_row=last)
cats = Reference(ws, min_col=1, min_row=4, max_row=last)
ch.add_data(data, titles_from_data=True); ch.set_categories(cats); ch.legend = None
ws.add_chart(ch, "H3")
ln = LineChart(); ln.title = "Shows and Active Promoters by Year"; ln.height = 8; ln.width = 18
d2 = Reference(ws, min_col=2, min_row=3, max_row=last)
d3 = Reference(ws, min_col=6, min_row=3, max_row=last)
ln.add_data(d2, titles_from_data=True); ln.add_data(d3, titles_from_data=True); ln.set_categories(cats)
ws.add_chart(ln, "H20")

# ================= Sheet 3: Top Promoters Overall =================
ws = wb.create_sheet("Top Promoters")
ws.sheet_view.showGridLines = False
ws["A1"] = "Top Promoters by US Gross, 2010-2025 (parents aggregated)"; ws["A1"].font = TITLE
hdr = ["Rank", "Promoter (parent)", "Tier", "Gross USD", "Market Share", "Shows", "Tickets Sold"]
ws.append([]); ws.append(hdr); style_header(ws, 3, len(hdr))
top = parent_gross.most_common(40)
for i, (p, g) in enumerate(top, 1):
    ws.append([i, p, tier_of(p), g, g / TOTAL, parent_shows[p], parent_tix[p]])
last = ws.max_row
for rr in range(4, last + 1):
    ws.cell(rr, 4).number_format = MONEY
    ws.cell(rr, 5).number_format = PCT
    ws.cell(rr, 6).number_format = NUM
    ws.cell(rr, 7).number_format = NUM
    for c in range(1, 8):
        ws.cell(rr, c).border = BORDER
autowidth(ws, [6, 38, 20, 18, 14, 10, 14])
ch = BarChart(); ch.type = "bar"; ch.title = "Top 15 Promoters by US Gross"; ch.height = 12; ch.width = 20
data = Reference(ws, min_col=4, min_row=3, max_row=3 + 15)
cats = Reference(ws, min_col=2, min_row=4, max_row=3 + 15)
ch.add_data(data, titles_from_data=True); ch.set_categories(cats); ch.legend = None
ws.add_chart(ch, "I3")

# ================= Sheet 4: Market Share by Year =================
ws = wb.create_sheet("Share by Year")
ws.sheet_view.showGridLines = False
ws["A1"] = "Market Share by Year (share of US gross)"; ws["A1"].font = TITLE
top_parents = [p for p, _ in parent_gross.most_common() if parent_seg[p] in ("Major", "Independent")][:10]
cols = ["Live Nation", "AEG Presents"] + [p for p in top_parents if p not in ("Live Nation", "AEG Presents")][:8]
hdr = ["Year"] + cols + ["All Other"]
ws.append([]); ws.append(hdr); style_header(ws, 3, len(hdr))
for y in YEARS:
    total_y = yr_gross[y]
    rowvals = [y]
    named = 0.0
    for p in cols:
        sh = py_gross[p].get(y, 0) / total_y if total_y else 0
        named += sh; rowvals.append(sh)
    rowvals.append(max(0.0, 1 - named))
    ws.append(rowvals)
last = ws.max_row
for rr in range(4, last + 1):
    for c in range(2, len(hdr) + 1):
        ws.cell(rr, c).number_format = PCT
autowidth(ws, [8] + [15] * (len(hdr) - 1))
ch = BarChart(); ch.type = "col"; ch.grouping = "percentStacked"; ch.overlap = 100
ch.title = "Promoter Market Share by Year (100% stacked)"; ch.height = 11; ch.width = 24
data = Reference(ws, min_col=2, min_row=3, max_col=len(hdr), max_row=last)
cats = Reference(ws, min_col=1, min_row=4, max_row=last)
ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
ws.add_chart(ch, "A20")

# ================= Sheet 5: Tier Summary by Year =================
ws = wb.create_sheet("Tier Summary")
ws.sheet_view.showGridLines = False
ws["A1"] = "Market Structure by Tier and Year (share of US gross)"; ws["A1"].font = TITLE
tiers = ["Tier 1 (Majors)", "Tier 2 (Large indie)", "Tier 3 (Small indie)", "Venue / In-House", "Non-music content"]
hdr = ["Year"] + tiers
ws.append([]); ws.append(hdr); style_header(ws, 3, len(hdr))
for y in YEARS:
    total_y = yr_gross[y]
    ws.append([y] + [tier_year[t].get(y, 0) / total_y if total_y else 0 for t in tiers])
last = ws.max_row
for rr in range(4, last + 1):
    for c in range(2, len(hdr) + 1):
        ws.cell(rr, c).number_format = PCT
autowidth(ws, [8] + [20] * len(tiers))
ch = AreaChart(); ch.grouping = "percentStacked"; ch.title = "Market Structure Over Time (Tier 1 vs Independents vs In-House)"
ch.height = 11; ch.width = 24
data = Reference(ws, min_col=2, min_row=3, max_col=len(hdr), max_row=last)
cats = Reference(ws, min_col=1, min_row=4, max_row=last)
ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
ws.add_chart(ch, "A20")
# absolute gross table alongside
ws.cell(1, 9, "Gross USD by Tier and Year").font = BOLD
ws.cell(3, 9, "Year").font = HDR; ws.cell(3, 9).fill = HDR_FILL
for j, t in enumerate(tiers, 10):
    cell = ws.cell(3, j, t); cell.font = HDR; cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
for i, y in enumerate(YEARS, 4):
    ws.cell(i, 9, y)
    for j, t in enumerate(tiers, 10):
        c = ws.cell(i, j, tier_year[t].get(y, 0)); c.number_format = MONEY

# ================= Sheet 6: Tier 2/3 Fragmentation =================
ws = wb.create_sheet("Fragmentation")
ws.sheet_view.showGridLines = False
ws["A1"] = "Fragmentation of the Independent (Tier 2 + Tier 3) Segment"; ws["A1"].font = TITLE
ws["A2"] = ("HHI = Herfindahl-Hirschman Index within the independent segment (0-10000). "
            "Lower = more fragmented. Top-10 share = % of independent gross held by the 10 largest independents that year.")
ws["A2"].font = SUB; ws["A2"].alignment = Alignment(wrap_text=True); ws.row_dimensions[2].height = 28
hdr = ["Year", "Active Independents", "Independent Gross USD", "HHI (independent segment)", "Top-10 Independent Share"]
HROW = 4
for j, h in enumerate(hdr, 1):
    ws.cell(HROW, j, h)
style_header(ws, HROW, len(hdr))
for i, y in enumerate(YEARS, HROW + 1):
    fy = frag[y]
    ws.cell(i, 1, y)
    ws.cell(i, 2, fy["active"]).number_format = NUM
    ws.cell(i, 3, fy["gross"]).number_format = MONEY
    ws.cell(i, 4, round(fy["hhi"])).number_format = NUM
    ws.cell(i, 5, fy["top10"] / 100).number_format = PCT
last = HROW + len(YEARS)
autowidth(ws, [8, 20, 22, 24, 24])
cats = Reference(ws, min_col=1, min_row=HROW + 1, max_row=last)
ln = LineChart(); ln.title = "Active Independent Promoters by Year"; ln.height = 9; ln.width = 18
ln.add_data(Reference(ws, min_col=2, min_row=HROW, max_row=last), titles_from_data=True)
ln.set_categories(cats); ln.legend = None
ws.add_chart(ln, "G4")
ln2 = LineChart(); ln2.title = "Concentration of Independent Segment (HHI)"; ln2.height = 9; ln2.width = 18
ln2.add_data(Reference(ws, min_col=4, min_row=HROW, max_row=last), titles_from_data=True)
ln2.set_categories(cats); ln2.legend = None
ws.add_chart(ln2, "G22")

# ================= Sheet 7: Tier 2 roster =================
ws = wb.create_sheet("Tier 2 Independents")
ws.sheet_view.showGridLines = False
ws["A1"] = f"Tier 2 Large Independents (>= ${TIER2_CUTOFF/1e6:.0f}M lifetime US gross)"; ws["A1"].font = TITLE
hdr = ["Rank", "Promoter (parent)", "Gross USD", "Share of US", "Share of Independent Seg.", "Shows", "Tickets"]
ws.append([]); ws.append(hdr); style_header(ws, 3, len(hdr))
indie_total = sum(parent_gross[p] for p in indie_parents)
t2 = sorted([(p, parent_gross[p]) for p in indie_parents if parent_gross[p] >= TIER2_CUTOFF], key=lambda x: -x[1])
for i, (p, g) in enumerate(t2, 1):
    ws.append([i, p, g, g / TOTAL, g / indie_total, parent_shows[p], parent_tix[p]])
last = ws.max_row
for rr in range(4, last + 1):
    ws.cell(rr, 3).number_format = MONEY
    ws.cell(rr, 4).number_format = PCT
    ws.cell(rr, 5).number_format = PCT
    ws.cell(rr, 6).number_format = NUM
    ws.cell(rr, 7).number_format = NUM
autowidth(ws, [6, 38, 18, 12, 22, 10, 14])
ws.cell(last + 2, 2, f"Tier 2 count: {len(t2)}   |   Tier 3 (smaller independents) count: {len(indie_parents)-len(t2):,}").font = BOLD

# ================= Sheet 8: Mapping / transparency =================
ws = wb.create_sheet("Promoter Mapping")
ws.sheet_view.showGridLines = False
ws["A1"] = "Promoter Roll-up Rules (for audit)"; ws["A1"].font = TITLE
r = 3
ws.cell(r, 1, "Live Nation includes brand patterns:").font = BOLD; r += 1
ws.cell(r, 1, ", ".join(LN_BRANDS)); r += 2
ws.cell(r, 1, "AEG Presents includes brand patterns:").font = BOLD; r += 1
ws.cell(r, 1, ", ".join(AEG_BRANDS)); r += 2
ws.cell(r, 1, "Non-music content owners (separate bucket):").font = BOLD; r += 1
ws.cell(r, 1, ", ".join(sorted({n for _, n in NONMUSIC}))); r += 2
ws.cell(r, 1, "Independent variant merges (canonical names):").font = BOLD; r += 1
for _, name in INDIE_CANON:
    pass
seen = []
for _, name in INDIE_CANON:
    if name not in seen:
        seen.append(name)
ws.cell(r, 1, ", ".join(seen)); r += 2
for rr in range(3, r):
    ws.cell(rr, 1).alignment = Alignment(wrap_text=True, vertical="top")
autowidth(ws, [130])

wb.save(OUT)
print(f"Saved {OUT} with {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}")
