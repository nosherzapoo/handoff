#!/usr/bin/env python3
"""
Weekly Prediction-Market Notional Volume — self-contained scraper + workbook builder.

One command turns the live Dune query "Weekly Prediction Market Notional Volume"
(https://dune.com/queries/5753743) into the `Weekly PM Notional Value.xlsx`
workbook: a `raw` sheet (one row per week x platform) and a `pivot` sheet
(platform x week, driven by SUMIFS formulas, plus Total row/column).

    STEP 1  scrape Dune query 5753743  ->  dune_5753743.csv
    STEP 2  build the two-sheet workbook  ->  Weekly PM Notional Value.xlsx

Typical use
-----------
    # Full run: scrape Dune, then build the workbook
    python3 build_pm_notional_workbook.py

    # Rebuild the workbook from an already-scraped CSV (no browser needed)
    python3 build_pm_notional_workbook.py --from-csv dune_5753743.csv

    # Custom output name
    python3 build_pm_notional_workbook.py --out "Weekly PM Notional Value (8th July).xlsx"

Why a real browser is needed for the scrape
--------------------------------------------
Dune is JavaScript-rendered behind Cloudflare. WebFetch / curl / the /csv
endpoint all return a Cloudflare challenge or an empty loading skeleton, and
api.dune.com needs a paid API key. The only reliable no-key path is to drive a
real Chrome via Playwright, let the Cloudflare clearance cookie persist in a
profile directory, read the on-screen results table, and page through it. That
is exactly what `scrape_dune()` does. The workbook builder itself needs no
browser, so `--from-csv` lets you (re)build offline.

Dependencies:  see requirements.txt   (openpyxl always; playwright + a real
Chrome only for the live scrape).
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import sys
import time
from datetime import datetime
from pathlib import Path

# --------------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------------- #

DUNE_QUERY_ID = 5753743
DUNE_URL = f"https://dune.com/queries/{DUNE_QUERY_ID}"

HERE = Path(__file__).resolve().parent
DEFAULT_CSV = HERE / "dune_5753743.csv"
DEFAULT_XLSX = HERE / "Weekly PM Notional Value.xlsx"

# Persistent Chrome profile: the Cloudflare clearance cookie is cached here so
# the second and subsequent runs clear the challenge in a couple of seconds.
PROFILE_DIR = "/tmp/dune_profile"

# The three columns the Dune query returns, in order.
COL_WEEK = "week"
COL_PLATFORM = "platform"
COL_NOTIONAL = "Notional USD Volume"

ROWS_PER_PAGE = 25          # Dune renders 25 rows per results page
NUM_FMT = "#,##0.00"        # notional dollar columns
DATE_FMT = "yyyy-mm-dd"     # week columns


# --------------------------------------------------------------------------- #
# STEP 1 — scrape Dune query 5753743
# --------------------------------------------------------------------------- #

# Read the largest on-screen <table> (Dune renders several) as {headers, rows}.
_READ_TABLE_JS = r"""
() => {
  const tables = Array.from(document.querySelectorAll('table'));
  if (!tables.length) return null;
  let best = tables[0], bestN = -1;
  for (const t of tables) {
    const n = t.querySelectorAll('tbody tr').length;
    if (n > bestN) { best = t; bestN = n; }
  }
  const headers = Array.from(best.querySelectorAll('thead th')).map(th => th.innerText.trim());
  const rows = Array.from(best.querySelectorAll('tbody tr')).map(tr =>
    Array.from(tr.querySelectorAll('td')).map(td => td.innerText.trim()));
  return { headers, rows };
}
"""

# The footer text like "698 rows" — tells us how many pages to page through.
_FOOTER_INFO_JS = r"""
() => {
  for (const s of Array.from(document.querySelectorAll('span'))) {
    const m = (s.innerText || '').trim().match(/^([\d,]+)\s+rows?$/i);
    if (m) return m[0];
  }
  return null;
}
"""

# Drive the page-number input directly (the Next button is unreliable). Uses the
# React-friendly value setter so React notices the change, then dispatches Enter.
_GOTO_PAGE_JS = r"""
(targetPage) => {
  const pageInput = Array.from(document.querySelectorAll('input')).find(i =>
    i.getAttribute('aria-label') !== 'Search' && i.type === 'text' && /^\d+$/.test(i.value || ''));
  if (!pageInput) return 'no-page-input';
  const setter = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value').set;
  setter.call(pageInput, String(targetPage));
  pageInput.dispatchEvent(new Event('input', { bubbles: true }));
  pageInput.dispatchEvent(new Event('change', { bubbles: true }));
  pageInput.dispatchEvent(new KeyboardEvent('keydown', { key: 'Enter', code: 'Enter', bubbles: true }));
  return 'ok';
}
"""

# The first row's text — used to detect that a page turn actually landed.
_FIRST_ROW_JS = r"""
() => {
  const tables = Array.from(document.querySelectorAll('table'));
  let best = null, bestN = -1;
  for (const t of tables) {
    const n = t.querySelectorAll('tbody tr').length;
    if (n > bestN) { best = t; bestN = n; }
  }
  const first = best && best.querySelector('tbody tr');
  return first ? first.innerText : null;
}
"""

_CLICK_BTN_JS = r"""
(label) => {
  const b = Array.from(document.querySelectorAll('button')).find(x => (x.innerText || '').trim() === label);
  if (b) { b.click(); return true; }
  return false;
}
"""

_BUTTON_TEXTS_JS = "() => Array.from(document.querySelectorAll('button')).map(b => (b.innerText||'').trim()).filter(Boolean)"

# The run/freshness control: a button whose text is a short age like "3h", "1d".
_CLICK_FRESHNESS_JS = r"""
() => {
  const b = Array.from(document.querySelectorAll('button')).find(x => /^\d+\s*[smhd]$/i.test((x.innerText || '').trim()));
  if (b) { b.click(); return true; }
  return false;
}
"""


def scrape_dune(csv_path: Path, headless: bool = False) -> list[dict]:
    """Scrape all rows of Dune query 5753743 and write them to ``csv_path``.

    Returns the parsed rows as a list of dicts with keys week/platform/notional.
    Requires `playwright` + `playwright-stealth` and a real Chrome
    (`playwright install chrome`, or the system Chrome the `channel="chrome"`
    launch picks up).
    """
    try:
        from playwright.sync_api import sync_playwright
        from playwright_stealth import Stealth
    except ImportError as exc:  # pragma: no cover - dependency hint
        sys.exit(
            "Live scrape needs playwright + playwright-stealth.\n"
            "  pip install -r requirements.txt && playwright install chrome\n"
            f"(import error: {exc})\n"
            "If you already have a dune_5753743.csv, rebuild offline with --from-csv."
        )

    print(f"[scrape] opening {DUNE_URL}")
    with Stealth().use_sync(sync_playwright()) as p:
        ctx = p.chromium.launch_persistent_context(
            PROFILE_DIR,
            channel="chrome",
            headless=headless,
            args=["--disable-blink-features=AutomationControlled", "--disk-cache-size=1"],
            viewport={"width": 1400, "height": 900},
        )
        page = ctx.pages[0] if ctx.pages else ctx.new_page()
        page.goto(DUNE_URL, wait_until="domcontentloaded")

        # 1. Wait out the Cloudflare challenge (fast once the cookie is cached).
        print("[scrape] waiting for Cloudflare...")
        for _ in range(120):
            title = page.title()
            if title and not any(x in title for x in ("Cloudflare", "Attention", "Just a moment")):
                print(f"[scrape] Cloudflare cleared — {title!r}")
                break
            time.sleep(1)
        else:
            ctx.close()
            sys.exit("[scrape] Cloudflare never cleared; aborting.")

        time.sleep(3)
        if page.evaluate(_CLICK_BTN_JS, "Accept"):
            print("[scrape] dismissed cookie banner")
            time.sleep(1)

        # 2. If results aren't cached, trigger a run and wait for the table.
        def n_rows() -> int:
            return page.evaluate("() => document.querySelectorAll('table tbody tr').length")

        if n_rows() == 0:
            btns = page.evaluate(_BUTTON_TEXTS_JS)
            print(f"[scrape] no cached table; buttons = {btns}")
            if any(b == "Running" for b in btns):
                print("[scrape] query already running, waiting...")
            elif page.evaluate(_CLICK_BTN_JS, "Run"):
                print("[scrape] clicked Run")
            elif page.evaluate(_CLICK_FRESHNESS_JS):
                print("[scrape] clicked freshness/run button")
            waited = 0
            while waited < 8 * 60:
                if n_rows() > 0:
                    print(f"[scrape] results rendered after ~{waited}s")
                    break
                time.sleep(3)
                waited += 3
            else:
                ctx.close()
                sys.exit("[scrape] results never rendered within 8 min; aborting.")

        # 3. Read page 1, work out the page count from the footer.
        data = page.evaluate(_READ_TABLE_JS)
        headers = data["headers"]
        all_rows = list(data["rows"])
        print(f"[scrape] headers = {headers}; {len(all_rows)} rows on page 1")

        footer = page.evaluate(_FOOTER_INFO_JS)
        total_rows = None
        if footer:
            m = re.search(r"([\d,]+)\s+rows?", footer)
            if m:
                total_rows = int(m.group(1).replace(",", ""))
        per_page = len(all_rows) or ROWS_PER_PAGE
        total_pages = math.ceil(total_rows / per_page) if total_rows else 1
        print(f"[scrape] footer={footer!r} -> total_rows={total_rows}, total_pages={total_pages}")

        # 4. Page through the rest, confirming each turn actually landed.
        for page_num in range(2, total_pages + 1):
            prev_first = page.evaluate(_FIRST_ROW_JS)
            if page.evaluate(_GOTO_PAGE_JS, page_num) != "ok":
                print(f"[scrape] page {page_num}: no page input, stopping")
                break
            changed = _wait_for_page_turn(page, prev_first)
            if not changed:  # one retry
                page.evaluate(_GOTO_PAGE_JS, page_num)
                changed = _wait_for_page_turn(page, prev_first, tries=40)
            if not changed:
                print(f"[scrape] page {page_num} never changed, stopping")
                break
            page_data = page.evaluate(_READ_TABLE_JS)
            all_rows.extend(page_data["rows"])
            print(f"[scrape] page {page_num}/{total_pages}: {len(page_data['rows'])} rows")

        ctx.close()

    if total_rows and len(all_rows) != total_rows:
        print(f"[scrape] WARNING: collected {len(all_rows)} rows, footer said {total_rows}")

    # 5. Normalise + persist to CSV, and return the parsed records.
    records = _rows_to_records(headers, all_rows)
    _write_csv(records, csv_path)
    print(f"[scrape] wrote {len(records)} rows -> {csv_path}")
    return records


def _wait_for_page_turn(page, prev_first: str, tries: int = 60) -> bool:
    for _ in range(tries):
        cur = page.evaluate(_FIRST_ROW_JS)
        if cur and cur != prev_first:
            return True
        time.sleep(0.5)
    return False


def _rows_to_records(headers: list[str], rows: list[list[str]]) -> list[dict]:
    """Map raw scraped cells to {week, platform, notional} records."""
    idx = {h: i for i, h in enumerate(headers)}
    # Be tolerant of small header changes: fall back to positional order.
    wi = idx.get(COL_WEEK, 0)
    pi = idx.get(COL_PLATFORM, 1)
    ni = idx.get(COL_NOTIONAL, 2)
    out = []
    for r in rows:
        if len(r) <= max(wi, pi, ni):
            continue
        out.append(
            {
                COL_WEEK: _parse_week(r[wi]),
                COL_PLATFORM: r[pi].strip(),
                COL_NOTIONAL: _parse_number(r[ni]),
            }
        )
    return out


def _parse_week(s: str) -> datetime:
    s = s.strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    # Last resort: take the leading date portion.
    return datetime.strptime(s[:10], "%Y-%m-%d")


def _parse_number(s: str) -> float:
    # Table viz may apply currency formatting -> "$3,705,896,611.1"; strip it.
    cleaned = re.sub(r"[^0-9.\-]", "", s)
    return float(cleaned) if cleaned not in ("", "-", ".") else 0.0


# --------------------------------------------------------------------------- #
# CSV helpers
# --------------------------------------------------------------------------- #

def _write_csv(records: list[dict], path: Path) -> None:
    with path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow([COL_WEEK, COL_PLATFORM, COL_NOTIONAL])
        for r in records:
            week = r[COL_WEEK]
            week_s = week.strftime("%Y-%m-%d %H:%M:%S") if isinstance(week, datetime) else str(week)
            w.writerow([week_s, r[COL_PLATFORM], r[COL_NOTIONAL]])


def read_csv(path: Path) -> list[dict]:
    """Read a previously-scraped dune_5753743.csv into records."""
    records = []
    with path.open() as f:
        for row in csv.DictReader(f):
            records.append(
                {
                    COL_WEEK: _parse_week(row[COL_WEEK]),
                    COL_PLATFORM: row[COL_PLATFORM].strip(),
                    COL_NOTIONAL: float(row[COL_NOTIONAL]),
                }
            )
    return records


# --------------------------------------------------------------------------- #
# STEP 2 — build the two-sheet workbook
# --------------------------------------------------------------------------- #

def build_workbook(records: list[dict], out_path: Path) -> None:
    """Build the `raw` + `pivot` workbook from records and save to out_path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    # ---- raw sheet: week ascending, platform alphabetical (matches Dune) ---- #
    rows = sorted(records, key=lambda r: (r[COL_WEEK], r[COL_PLATFORM]))
    weeks = sorted({r[COL_WEEK] for r in records})
    # Platforms ordered by total notional, descending (drives pivot row order).
    totals: dict[str, float] = {}
    for r in records:
        totals[r[COL_PLATFORM]] = totals.get(r[COL_PLATFORM], 0.0) + r[COL_NOTIONAL]
    platforms = sorted(totals, key=lambda p: -totals[p])

    wb = Workbook()
    ws = wb.active
    ws.title = "raw"
    bold = Font(bold=True)

    ws.append([COL_WEEK, COL_PLATFORM, COL_NOTIONAL])
    for c in range(1, 4):
        ws.cell(1, c).font = bold
    for r in rows:
        ws.append([r[COL_WEEK], r[COL_PLATFORM], r[COL_NOTIONAL]])
    n_data = len(rows)
    last = n_data + 1  # last data row (row 1 is the header)
    for r in range(2, last + 1):
        ws.cell(r, 1).number_format = DATE_FMT
        ws.cell(r, 3).number_format = NUM_FMT
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 20

    # ---- pivot sheet: platform x week via SUMIFS, + Total row/column -------- #
    pv = wb.create_sheet("pivot")
    pv.cell(1, 1, "platform \\ week").font = bold
    for j, wk in enumerate(weeks):
        c = pv.cell(1, 2 + j, wk)
        c.font = bold
        c.number_format = DATE_FMT
    total_col = 2 + len(weeks)
    pv.cell(1, total_col, "Total").font = bold

    c_first = get_column_letter(2)
    c_last = get_column_letter(1 + len(weeks))
    raw_c = f"raw!$C$2:$C${last}"
    raw_a = f"raw!$A$2:$A${last}"
    raw_b = f"raw!$B$2:$B${last}"

    for i, plat in enumerate(platforms):
        row = 2 + i
        pv.cell(row, 1, plat)
        for j in range(len(weeks)):
            col = 2 + j
            wl = get_column_letter(col)
            cell = pv.cell(
                row, col,
                f"=SUMIFS({raw_c},{raw_a},{wl}$1,{raw_b},$A{row})",
            )
            cell.number_format = NUM_FMT
        tcell = pv.cell(row, total_col, f"=SUM({c_first}{row}:{c_last}{row})")
        tcell.number_format = NUM_FMT

    total_row = 2 + len(platforms)
    pv.cell(total_row, 1, "Total").font = bold
    for col in range(2, total_col):  # per-week totals: sum the platforms above
        wl = get_column_letter(col)
        cell = pv.cell(total_row, col, f"=SUM({wl}2:{wl}{total_row - 1})")
        cell.number_format = NUM_FMT
        cell.font = bold
    # Grand-total corner: sum the Total row across the week columns.
    corner = pv.cell(
        total_row, total_col,
        f"=SUM({c_first}{total_row}:{c_last}{total_row})",
    )
    corner.number_format = NUM_FMT
    corner.font = bold

    pv.freeze_panes = "B2"
    pv.column_dimensions["A"].width = 16

    wb.save(out_path)
    print(
        f"[build] {out_path}  —  raw: {n_data} rows, "
        f"pivot: {len(platforms)} platforms x {len(weeks)} weeks"
    )


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(
        description="Scrape Dune query 5753743 and build the Weekly PM Notional Value workbook.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument(
        "--from-csv", metavar="PATH", type=Path, default=None,
        help="Skip the browser scrape and build from this already-scraped CSV.",
    )
    ap.add_argument(
        "--csv", metavar="PATH", type=Path, default=DEFAULT_CSV,
        help=f"Where to write/read the scraped CSV (default: {DEFAULT_CSV.name}).",
    )
    ap.add_argument(
        "--out", metavar="PATH", type=Path, default=DEFAULT_XLSX,
        help=f"Output workbook path (default: {DEFAULT_XLSX.name}).",
    )
    ap.add_argument(
        "--headless", action="store_true",
        help="Run Chrome headless during the scrape (may be flagged by Cloudflare).",
    )
    ap.add_argument(
        "--no-build", action="store_true",
        help="Only scrape to CSV; do not build the workbook.",
    )
    args = ap.parse_args(argv)

    if args.from_csv:
        if not args.from_csv.exists():
            ap.error(f"--from-csv file not found: {args.from_csv}")
        print(f"[read] loading {args.from_csv}")
        records = read_csv(args.from_csv)
    else:
        records = scrape_dune(args.csv, headless=args.headless)

    if not records:
        sys.exit("No records — nothing to build.")

    if not args.no_build:
        build_workbook(records, args.out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
