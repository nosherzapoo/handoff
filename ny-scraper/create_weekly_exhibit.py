#!/usr/bin/env python3
"""
Creates a weekly exhibit Excel file showing the most recent 5 weeks of NY gaming data.
Layout: Week | Handle / GGR / Hold per operator + Statewide
Each week has two rows: actual data row + yy increase row with signed % / bps.
Alternating white / light-gray band per week pair (data row + yy row share the same band).
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ── Colours ───────────────────────────────────────────────────────────────────
NAVY        = "1F3864"
ALT_GRAY    = "D9D9D9"   # alternating band colour
GREEN       = "00B050"
RED         = "FF0000"
BLACK       = "000000"
GRAY_TEXT   = "595959"
WHITE       = "FFFFFF"

HEADER_FILL   = PatternFill(start_color=NAVY,     end_color=NAVY,     fill_type="solid")
WHITE_FILL    = PatternFill(start_color=WHITE,    end_color=WHITE,    fill_type="solid")
GRAY_FILL     = PatternFill(start_color=ALT_GRAY, end_color=ALT_GRAY, fill_type="solid")

# ── Fonts ─────────────────────────────────────────────────────────────────────
HEADER_FONT    = Font(color=WHITE,     bold=True,   size=10, name="Arial")
DATA_FONT      = Font(color=BLACK,                  size=10, name="Arial")
BOLD_DATA_FONT = Font(color=BLACK,     bold=True,   size=10, name="Arial")
POS_FONT       = Font(color=GREEN,                  size=9,  name="Arial")
NEG_FONT       = Font(color=RED,                    size=9,  name="Arial")
YOY_LABEL_FONT = Font(color=GRAY_TEXT, italic=True, size=9,  name="Arial")

# ── Border sides ──────────────────────────────────────────────────────────────
# Only two rules:
#   1. Medium gray vertical line on the outer edges of every brand group
#   2. Medium gray horizontal line under the sub-header row (row 2)
_MG = Side(style="medium", color="595959")
_NO = Side(style=None)

def _bdr(left=_NO, right=_NO, top=_NO, bottom=_NO):
    return Border(left=left, right=right, top=top, bottom=bottom)

def cell_border(col: int, is_header_bottom: bool = False) -> Border:
    """
    col is 1-based.
    Col 1  = Week (single-column group → left + right).
    Col 2+ = brand data columns in groups of 3.
      left  border if (col - 2) % 3 == 0  (first of group)
      right border if (col - 2) % 3 == 2  (last of group)
    """
    if col == 1:
        left_side  = _MG
        right_side = _MG
    else:
        offset     = col - 2
        left_side  = _MG if offset % 3 == 0 else _NO
        right_side = _MG if offset % 3 == 2 else _NO
    bottom_side = _MG if is_header_bottom else _NO
    return _bdr(left=left_side, right=right_side, bottom=bottom_side)

# ── Brand display names ───────────────────────────────────────────────────────
BRAND_DISPLAY = {
    'Bally Bet':                'Bally Bet',
    'BetMGM':                   'BetMGM',
    'Caesars Sport Book':       'Caesars',
    'DraftKings Sport Book':    'DraftKings',
    'ESPN Bet':                 'ESPN Bet',
    'Wynn Interactive':         'Wynn',
    'FanDuel':                  'FanDuel',
    'Fanatics':                 'Fanatics',
    'Resorts World Bet':        'Resorts World',
    'Rush Street Interactive':  'Rush Street',
}

def shorten(brand: str) -> str:
    return BRAND_DISPLAY.get(brand, brand)

# ── YoY date lookup ───────────────────────────────────────────────────────────
def find_yoy_date(current_date, all_dates):
    target = current_date - pd.DateOffset(days=364)
    window = [d for d in all_dates if abs((d - target).days) <= 7]
    if not window:
        return None
    return min(window, key=lambda x: abs((x - target).days))

# ── Value formatters ──────────────────────────────────────────────────────────
def fmt_yoy_pct(val):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    sign = '+' if val > 0 else ''
    return f"{sign}{val * 100:.0f}%"

def fmt_yoy_bps(diff):
    if diff is None or (isinstance(diff, float) and np.isnan(diff)):
        return None
    bps  = diff * 10000
    sign = '+' if bps > 0 else ''
    return f"{sign}{bps:.0f}bps"

# ── Main ──────────────────────────────────────────────────────────────────────
def create_weekly_exhibit(data_file: str = 'ny_gaming_data.csv',
                          output_file: str = 'ny_gaming_weekly_exhibit.xlsx',
                          num_weeks: int = 5) -> str:
    logger.info("📋 Creating weekly exhibit...")

    df = pd.read_csv(data_file)
    df['Date'] = pd.to_datetime(df['Date'])

    all_dates    = sorted(df['Date'].unique(), reverse=True)
    recent_dates = all_dates[:num_weeks]

    handle_piv = df.pivot_table(index='Date', columns='Brand', values='Handle', aggfunc='sum')
    ggr_piv    = df.pivot_table(index='Date', columns='Brand', values='GGR',    aggfunc='sum')
    handle_piv['Statewide'] = handle_piv.sum(axis=1)
    ggr_piv['Statewide']    = ggr_piv.sum(axis=1)
    hold_piv = ggr_piv / handle_piv.replace(0, np.nan)

    FEATURED_BRANDS = ['DraftKings Sport Book', 'FanDuel', 'BetMGM', 'Fanatics']
    available = set(df['Brand'].unique())
    brands    = [b for b in FEATURED_BRANDS if b in available]
    all_cols  = brands + ['Statewide']
    num_groups = len(all_cols)
    last_col   = 1 + num_groups * 3   # rightmost column index

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = 'Weekly Exhibit'
    ws.sheet_view.showGridLines = False   # turn off gridlines

    # ── Helper: return the right fill based on alternating band index ─────────
    def band_fill(week_idx: int) -> PatternFill:
        return WHITE_FILL if week_idx % 2 == 0 else GRAY_FILL

    # ── Helper: safe pivot lookup ─────────────────────────────────────────────
    def safe_get(piv, date, col):
        try:
            if date in piv.index and col in piv.columns:
                v = piv.at[date, col]
                return v if not (isinstance(v, float) and np.isnan(v)) else None
        except Exception:
            pass
        return None

    # ── Helper: build border for a data/yoy cell ─────────────────────────────
    # col_pos: 0-based position within the row (0 = Week column)
    # group positions: col 0 = Week; then groups start at col 1+
    def data_border(col_pos: int) -> Border:
        """
        col_pos is 0-based (0 = Week column, 1 = first brand col, etc.)
        last_col is 1-based column index == col_pos + 1
        A medium border goes on the RIGHT of the last column in each group
        (every 3rd data column and the Week column itself).
        """
        is_group_right = (col_pos == 0) or ((col_pos - 1) % 3 == 2)
        right = _MG if is_group_right else _TG
        return _bdr(left=_TG, right=right, top=_TG, bottom=_TG)

    def header_border(col_pos: int) -> Border:
        is_group_right = (col_pos == 0) or ((col_pos - 1) % 3 == 2)
        right = _MW if is_group_right else _TW
        return _bdr(left=_TW, right=right, top=_TW, bottom=_TW)

    # ── Row 1: brand / group headers ──────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    c = ws.cell(row=1, column=1, value="Week")
    c.fill      = HEADER_FILL
    c.font      = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border    = cell_border(col=1)

    for idx, col_name in enumerate(all_cols):
        sc    = 2 + idx * 3
        label = 'Statewide' if col_name == 'Statewide' else shorten(col_name)
        ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=sc + 2)
        # Value + style on the first cell of the merge
        c = ws.cell(row=1, column=sc, value=label)
        c.fill      = HEADER_FILL
        c.font      = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = _bdr(left=_MG)
        # Right border on the last cell of the merge
        ws.cell(row=1, column=sc + 2).border = _bdr(right=_MG)

    # ── Row 2: Handle / GGR / Hold sub-headers (light gray) ──────────────────
    ROW2_FONT = Font(color=BLACK, bold=True, size=10, name="Arial")
    ws.cell(row=2, column=1).fill      = GRAY_FILL
    ws.cell(row=2, column=1).font      = ROW2_FONT
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=2, column=1).border    = cell_border(col=1, is_header_bottom=True)

    for idx in range(num_groups):
        for j, metric in enumerate(['Handle', 'GGR', 'Hold']):
            col = 2 + idx * 3 + j
            c = ws.cell(row=2, column=col, value=metric)
            c.fill      = GRAY_FILL
            c.font      = ROW2_FONT
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = cell_border(col=col, is_header_bottom=True)

    # ── Data + YoY rows ───────────────────────────────────────────────────────
    cur_row = 3

    for week_idx, date in enumerate(recent_dates):
        fill      = band_fill(week_idx)
        date_str  = f"{date.month}/{date.day}/{date.year}"
        yoy_date  = find_yoy_date(date, all_dates)

        # ── Data row ──────────────────────────────────────────────────────────
        d = ws.cell(row=cur_row, column=1, value=date_str)
        d.fill      = fill
        d.font      = BOLD_DATA_FONT
        d.alignment = Alignment(horizontal='center', vertical='center')
        d.border    = cell_border(col=1)

        for idx, col_name in enumerate(all_cols):
            sc = 2 + idx * 3
            hv = safe_get(handle_piv, date, col_name)
            gv = safe_get(ggr_piv,    date, col_name)
            ho = safe_get(hold_piv,   date, col_name)

            h_cell = ws.cell(row=cur_row, column=sc)
            h_cell.value         = int(round(hv)) if hv is not None else None
            h_cell.number_format = '#,##0'
            h_cell.font          = DATA_FONT
            h_cell.fill          = fill
            h_cell.alignment     = Alignment(horizontal='center', vertical='center')
            h_cell.border        = cell_border(col=sc)

            g_cell = ws.cell(row=cur_row, column=sc + 1)
            g_cell.value         = int(round(gv)) if gv is not None else None
            g_cell.number_format = '#,##0'
            g_cell.font          = DATA_FONT
            g_cell.fill          = fill
            g_cell.alignment     = Alignment(horizontal='center', vertical='center')
            g_cell.border        = cell_border(col=sc + 1)

            hold_cell = ws.cell(row=cur_row, column=sc + 2)
            if ho is not None:
                hold_cell.value         = ho
                hold_cell.number_format = '0.0%'
            hold_cell.font      = DATA_FONT
            hold_cell.fill      = fill
            hold_cell.alignment = Alignment(horizontal='center', vertical='center')
            hold_cell.border    = cell_border(col=sc + 2)

        cur_row += 1

        # ── YoY row ───────────────────────────────────────────────────────────
        yl = ws.cell(row=cur_row, column=1, value="yy increase")
        yl.fill      = fill
        yl.font      = YOY_LABEL_FONT
        yl.alignment = Alignment(horizontal='center', vertical='center')
        yl.border    = cell_border(col=1)

        for idx, col_name in enumerate(all_cols):
            sc = 2 + idx * 3

            # Apply fill + border to all three cells upfront
            for j in range(3):
                c = ws.cell(row=cur_row, column=sc + j)
                c.fill      = fill
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border    = cell_border(col=sc + j)

            if yoy_date is not None:
                ch = safe_get(handle_piv, date,     col_name)
                ph = safe_get(handle_piv, yoy_date, col_name)
                cg = safe_get(ggr_piv,    date,     col_name)
                pg = safe_get(ggr_piv,    yoy_date, col_name)
                co = safe_get(hold_piv,   date,     col_name)
                po = safe_get(hold_piv,   yoy_date, col_name)

                handle_yoy = (ch - ph) / ph if (ch and ph and ph != 0) else None
                ggr_yoy    = (cg - pg) / pg if (cg and pg and pg != 0) else None
                hold_diff  = (co - po)      if (co is not None and po is not None) else None

                h_yoy = ws.cell(row=cur_row, column=sc)
                txt = fmt_yoy_pct(handle_yoy)
                if txt:
                    h_yoy.value = txt
                    h_yoy.font  = POS_FONT if handle_yoy >= 0 else NEG_FONT

                g_yoy = ws.cell(row=cur_row, column=sc + 1)
                txt = fmt_yoy_pct(ggr_yoy)
                if txt:
                    g_yoy.value = txt
                    g_yoy.font  = POS_FONT if ggr_yoy >= 0 else NEG_FONT

                ho_yoy = ws.cell(row=cur_row, column=sc + 2)
                txt = fmt_yoy_bps(hold_diff)
                if txt:
                    ho_yoy.value = txt
                    ho_yoy.font  = POS_FONT if hold_diff >= 0 else NEG_FONT

        cur_row += 1

    # ── Bottom border under the last data row (row 12) ───────────────────────
    last_data_row = cur_row - 1
    for col in range(1, last_col + 1):
        c = ws.cell(row=last_data_row, column=col)
        existing = c.border
        c.border = Border(
            left   = existing.left,
            right  = existing.right,
            top    = existing.top,
            bottom = _MG,
        )

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 13
    for idx in range(num_groups):
        for j, metric in enumerate(['Handle', 'GGR', 'Hold']):
            col_letter = get_column_letter(2 + idx * 3 + j)
            ws.column_dimensions[col_letter].width = 15 if metric in ('Handle', 'GGR') else 9

    # ── Row heights ───────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    for r in range(3, cur_row):
        ws.row_dimensions[r].height = 16

    # ── Freeze header rows (no separate apply_borders call needed) ───────────
    ws.freeze_panes = 'B3'

    wb.save(output_file)
    logger.info(f"✅ Weekly exhibit saved: {output_file}")
    return output_file


if __name__ == "__main__":
    create_weekly_exhibit()
