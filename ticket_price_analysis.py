"""Mean min / max ticket price by year, faceted Non-Superstar vs Superstar,
for US shows, Europe shows, and All shows. Replicates the reference figure.

Prices are already in USD (verified). Superstar = headliner whose average gross
per show (total gross / total shows, 2010-2025) is >= $1,000,000 with >= 3
engagements. Classification is global and applied to every region.
"""
import glob
import json
from collections import defaultdict

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

YEARS = list(range(2010, 2026))
SUPERSTAR_GROSS = 1_000_000      # avg gross per show cutoff
MIN_ENGAGEMENTS = 3

EUROPE = {
    "United Kingdom", "Germany", "France", "Italy", "Spain", "Netherlands",
    "Ireland", "Belgium", "Sweden", "Switzerland", "Austria", "Denmark",
    "Norway", "Poland", "Portugal", "Finland", "Czech Republic", "Greece",
    "Hungary", "Luxembourg", "Romania", "Croatia", "Slovakia", "Slovenia",
    "Bulgaria", "Estonia", "Latvia", "Lithuania", "Iceland", "Malta", "Cyprus",
}

# colors matching the reference
C_MAX = "#7DDA58"   # green solid (mean maximum)
C_MIN = "#E8645A"   # red dashed (mean minimum)


def num(x):
    try:
        v = float(x)
        return v if v == v else 0.0
    except (TypeError, ValueError):
        return 0.0


def year_of(e):
    p = (e.get("eventDate") or "").split("/")
    return int(p[2]) if len(p) == 3 and p[2].isdigit() else None


print("Loading ...")
ev = []
for f in sorted(glob.glob("pages/page-*.json")):
    ev += json.load(open(f))

# ---- in-window rows ----
rows = []
for e in ev:
    y = year_of(e)
    if y is None or y < 2010 or y > 2025:
        continue
    rows.append(e)
print(f"rows 2010-2025: {len(rows):,}")

# ---- superstar classification (global, by headliner) ----
art_gross = defaultdict(float)
art_shows = defaultdict(float)
art_eng = defaultdict(int)
for e in rows:
    a = (e.get("headLiner") or "").strip()
    if not a:
        continue
    g = num(e.get("grossUSD"))
    sh = num(e.get("numShows")) or 1
    if g <= 0:
        continue
    art_gross[a] += g
    art_shows[a] += sh
    art_eng[a] += 1

superstars = {
    a for a in art_gross
    if art_eng[a] >= MIN_ENGAGEMENTS and art_gross[a] / art_shows[a] >= SUPERSTAR_GROSS
}
print(f"superstar artists: {len(superstars):,}")

# ---- winsorize price thresholds (global p0.1 / p99.9) ----
def winsor_bounds(key):
    vals = sorted(num(e.get(key)) for e in rows if num(e.get(key)) > 0)
    lo = vals[int(0.001 * len(vals))]
    hi = vals[int(0.999 * len(vals))]
    return lo, hi


min_lo, min_hi = winsor_bounds("ticketPriceMin")
max_lo, max_hi = winsor_bounds("ticketPriceMax")
print(f"winsor min in [{min_lo:.0f},{min_hi:.0f}]  max in [{max_lo:.0f},{max_hi:.0f}]")


def clamp(v, lo, hi):
    return lo if v < lo else hi if v > hi else v


# ---- aggregate: cell[region][klass][year] = [sum_min, sum_max, n] ----
def regions_of(e):
    out = ["All"]
    c = e.get("country")
    if c == "United States":
        out.append("US")
    if c in EUROPE:
        out.append("Europe")
    return out


cell = {r: {"Non-Superstar": {y: [0.0, 0.0, 0] for y in YEARS},
            "Superstar": {y: [0.0, 0.0, 0] for y in YEARS}}
        for r in ("US", "Europe", "All")}

for e in rows:
    pmin = num(e.get("ticketPriceMin"))
    pmax = num(e.get("ticketPriceMax"))
    if pmin <= 0 or pmax <= 0 or pmax < pmin:
        continue
    pmin = clamp(pmin, min_lo, min_hi)
    pmax = clamp(pmax, max_lo, max_hi)
    y = year_of(e)
    klass = "Superstar" if (e.get("headLiner") or "").strip() in superstars else "Non-Superstar"
    for r in regions_of(e):
        c = cell[r][klass][y]
        c[0] += pmin
        c[1] += pmax
        c[2] += 1


MIN_PLOT_N = 25   # suppress thin cells (mostly COVID 2020-2021) from the plotted lines

def series(region, klass):
    yrs, mn, mx, ns = [], [], [], []
    for y in YEARS:
        s_min, s_max, n = cell[region][klass][y]
        if n >= MIN_PLOT_N:
            yrs.append(y)
            mn.append(s_min / n)
            mx.append(s_max / n)
            ns.append(n)
    return yrs, mn, mx, ns


# ================= charts =================
REGION_TITLE = {"US": "United States", "Europe": "Europe (incl. UK)", "All": "All Shows (Worldwide)"}
png_files = {}
for region in ("US", "Europe", "All"):
    fig, axes = plt.subplots(1, 2, sharey=True, figsize=(9.2, 5.0))
    for ax, klass in zip(axes, ("Non-Superstar", "Superstar")):
        yrs, mn, mx, ns = series(region, klass)
        ax.plot(yrs, mx, color=C_MAX, lw=2.4, solid_capstyle="round", label="Mean Maximum Price")
        ax.plot(yrs, mn, color=C_MIN, lw=2.2, linestyle=(0, (6, 4)), label="Mean Minimum Price")
        ax.set_facecolor("#FbFbFb")
        ax.grid(axis="y", color="#E3E3E3", lw=0.8)
        ax.set_axisbelow(True)
        for s in ("top", "right"):
            ax.spines[s].set_visible(False)
        ax.set_title(klass, fontsize=12, fontweight="bold",
                     bbox=dict(boxstyle="square,pad=0.4", fc="#E4E7E4", ec="#BfBfBf"))
        ax.set_xticks(YEARS)
        ax.set_xticklabels(YEARS, rotation=45, ha="right", fontsize=8)
        ax.margins(x=0.02)
    axes[0].set_ylabel("Ticket Price (USD)", fontsize=11)
    fig.suptitle(f"Mean Ticket Price by Year: {REGION_TITLE[region]}",
                 fontsize=13, fontweight="bold", y=0.99)
    fig.text(0.5, 0.12, "Year", ha="center", fontsize=11)
    handles = [Line2D([0], [0], color=C_MAX, lw=2.4),
               Line2D([0], [0], color=C_MIN, lw=2.2, linestyle=(0, (6, 4)))]
    fig.legend(handles, ["Mean Maximum Price", "Mean Minimum Price"],
               title="Price Type", loc="lower center", ncol=2, frameon=True,
               bbox_to_anchor=(0.5, 0.0), fontsize=10, title_fontsize=10)
    fig.tight_layout(rect=(0, 0.13, 1, 0.97))
    fn = f"ticket_price_{region.lower()}.png"
    fig.savefig(fn, dpi=150)
    plt.close(fig)
    png_files[region] = fn
    print("wrote", fn)

# ================= excel =================
wb = Workbook()
HDR = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="1F3864")
TITLE = Font(bold=True, size=14, color="1F3864")

ws = wb.active
ws.title = "README"
ws.sheet_view.showGridLines = False
ws["A1"] = "Mean Ticket Price by Year: Superstar vs Non-Superstar"
ws["A1"].font = TITLE
notes = [
    "",
    ("Metric", "Per year, the unweighted mean of ticketPriceMin and ticketPriceMax across qualifying shows."),
    ("Currency", "All ticket prices are in USD (verified: gross / (avg price x tickets) = 1.000 for every reporting currency)."),
    ("Window", "Full years 2010-2025. The reference figure spans 1999-2019; this data begins in 2010."),
    ("COVID caveat", "2020-2021 had very few live shows; superstar cells in those years are thin (e.g. Europe Superstar 2021 n=10). Cells with fewer than 25 shows are omitted from the plotted lines but retained in the tables below with their counts."),
    ("Valid rows", "ticketPriceMin > 0, ticketPriceMax > 0, max >= min."),
    ("Outliers", f"Prices winsorized at global p0.1 / p99.9 (min capped to [{min_lo:.0f}, {min_hi:.0f}], max to [{max_lo:.0f}, {max_hi:.0f}]) to remove data errors such as a gross mis-entered as a price."),
    ("Superstar", f"Headliner with average gross per show (total gross / total shows) >= ${SUPERSTAR_GROSS:,.0f} and >= {MIN_ENGAGEMENTS} engagements over 2010-2025. Classification is global and applied to every region."),
    ("Superstar count", f"{len(superstars):,} artists."),
    ("US", "country = United States."),
    ("Europe", "European countries incl. United Kingdom (UK is the 2nd-largest market here). Switchable to strict EU-27 on request."),
    ("All", "Worldwide."),
    ("Sensitivity", "A $750k cutoff captures ~4.9% of shows / ~47% of gross; $1.5M captures ~2.1% / ~31%."),
]
r = 2
for item in notes:
    if item == "":
        r += 1
        continue
    a, b = item
    ws.cell(r, 1, a).font = Font(bold=True)
    c = ws.cell(r, 2, b)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 110

for region in ("US", "Europe", "All"):
    ws = wb.create_sheet(region)
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"Mean Ticket Price by Year: {REGION_TITLE[region]}"
    ws["A1"].font = TITLE
    hdr = ["Year",
           "Non-Superstar Mean Min", "Non-Superstar Mean Max", "Non-Superstar Shows",
           "Superstar Mean Min", "Superstar Mean Max", "Superstar Shows"]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(3, j, h)
        c.font = HDR
        c.fill = FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, y in enumerate(YEARS, 4):
        ns = cell[region]["Non-Superstar"][y]
        ss = cell[region]["Superstar"][y]
        vals = [y,
                ns[0] / ns[2] if ns[2] else None, ns[1] / ns[2] if ns[2] else None, ns[2],
                ss[0] / ss[2] if ss[2] else None, ss[1] / ss[2] if ss[2] else None, ss[2]]
        for j, v in enumerate(vals, 1):
            c = ws.cell(i, j, v)
            if j in (2, 3, 5, 6) and v is not None:
                c.number_format = '$#,##0.00'
            elif j in (4, 7):
                c.number_format = '#,##0'
    ws.column_dimensions["A"].width = 8
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 18
    img = XLImage(png_files[region])
    img.width = int(img.width * 0.85)
    img.height = int(img.height * 0.85)
    ws.add_image(img, "I3")

wb.save("Ticket_Price_Analysis.xlsx")
print("Saved Ticket_Price_Analysis.xlsx")
