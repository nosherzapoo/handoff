#!/usr/bin/env python3
"""
Quarterly attendance-weighted ticket-price analysis, US, 2019+.

- Weighted by ticketsSold (big shows count more).
- Series: ticketPriceMin (get-in), ticketPriceAvg, ticketPriceMax.
- Spread ratio = weighted_max / weighted_min per quarter.
- Deflated to constant 2025 USD via CPI-U (CPIAUCSL, seasonally adjusted, FRED).
- Nominal + real, indexed to 100 at 2019Q1, plus standalone spread-ratio line.
"""
import csv
from datetime import datetime
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DIR = "/Users/nosherzapoo/Desktop/claude/pollstar"

# ---------------------------------------------------------------- CPI-U
cpi_m = {}
with open(f"{DIR}/cpi_u_cpiaucsl.csv") as f:
    for row in csv.DictReader(f):
        v = row["CPIAUCSL"].strip()
        if v:
            cpi_m[row["observation_date"][:7]] = float(v)
# Oct 2025 missing in source -> interpolate from Sep & Nov
if "2025-10" not in cpi_m:
    cpi_m["2025-10"] = (cpi_m["2025-09"] + cpi_m["2025-11"]) / 2

def q_of(ym_month):
    return (ym_month - 1) // 3 + 1

# quarterly CPI = mean of the 3 monthly values
cpi_q = {}
tmp = {}
for ym, v in cpi_m.items():
    y, m = int(ym[:4]), int(ym[5:7])
    tmp.setdefault((y, q_of(m)), []).append(v)
for (y, q), vals in tmp.items():
    cpi_q[f"{y}Q{q}"] = sum(vals) / len(vals)

# base = 2025 full-year average CPI  -> "constant 2025 USD"
cpi_2025_avg = np.mean([cpi_m[f"2025-{mm:02d}"] for mm in range(1, 13)])
print(f"CPI-U 2025 annual avg (base) = {cpi_2025_avg:.3f}")

# ---------------------------------------------------------------- load data
rows = []
with open(f"{DIR}/concerts.csv") as f:
    for row in csv.DictReader(f):
        if row["country"] != "United States":
            continue
        try:
            d = datetime.strptime(row["eventDate"], "%m/%d/%Y")
        except ValueError:
            continue
        if d.year < 2019 or d.year >= 2026:   # 2026+ = forward-dated announced shows, excluded
            continue
        try:
            ts  = float(row["ticketsSold"])
            pmn = float(row["ticketPriceMin"])
            pav = float(row["ticketPriceAvg"])
            pmx = float(row["ticketPriceMax"])
        except ValueError:
            continue
        if ts <= 0 or pmn <= 0 or pav <= 0 or pmx <= 0:
            continue
        if pmx < pmn:                          # data-entry inversion
            continue
        rows.append((f"{d.year}Q{(d.month-1)//3+1}", d.year, ts, pmn, pav, pmx))

df = pd.DataFrame(rows, columns=["quarter", "year", "ts", "pmin", "pavg", "pmax"])
print(f"US valid events 2019-2025: {len(df):,}")

# ---------------------------------------------------------------- winsorize price outliers
# ticketPriceMax carries known data-entry errors (a gross mis-keyed as a price).
# Clip each price field to its [p0.1, p99.9] over the analysis set before weighting.
for col in ["pmin", "pavg", "pmax"]:
    lo, hi = df[col].quantile([0.001, 0.999])
    n_clip = int((df[col] < lo).sum() + (df[col] > hi).sum())
    df[col] = df[col].clip(lo, hi)
    print(f"  winsorized {col}: [{lo:.2f}, {hi:.2f}]  clipped {n_clip} rows")

# ---------------------------------------------------------------- weighted quarterly aggregation
def wmean(g, col):
    return np.average(g[col], weights=g["ts"])

recs = []
for q, g in df.groupby("quarter"):
    recs.append({
        "quarter": q,
        "n_events": len(g),
        "tickets": g["ts"].sum(),
        "w_min_nom": wmean(g, "pmin"),
        "w_avg_nom": wmean(g, "pavg"),
        "w_max_nom": wmean(g, "pmax"),
    })
res = pd.DataFrame(recs).sort_values("quarter").reset_index(drop=True)

# real (2025 USD)
res["defl"] = res["quarter"].map(lambda q: cpi_2025_avg / cpi_q[q])
for s in ["min", "avg", "max"]:
    res[f"w_{s}_real"] = res[f"w_{s}_nom"] * res["defl"]

# spread ratio (unit-free -> identical nominal & real)
res["spread_nom"]  = res["w_max_nom"]  / res["w_min_nom"]
res["spread_real"] = res["w_max_real"] / res["w_min_real"]

# thin-sample flag (COVID collapse): normal US quarters ~7-10k events
res["thin"] = res["n_events"] < 2000

# index to 100 at 2019Q1
base = res[res["quarter"] == "2019Q1"].iloc[0]
for s in ["min", "avg", "max"]:
    res[f"idx_{s}_real"] = 100 * res[f"w_{s}_real"] / base[f"w_{s}_real"]
    res[f"idx_{s}_nom"]  = 100 * res[f"w_{s}_nom"]  / base[f"w_{s}_nom"]
res["idx_spread"] = 100 * res["spread_real"] / base["spread_real"]

pd.set_option("display.width", 200, "display.max_columns", 40)
print("\n=== Quarterly (real 2025$) ===")
print(res[["quarter","n_events","w_min_real","w_avg_real","w_max_real","spread_real","thin"]]
      .to_string(index=False, float_format=lambda x: f"{x:,.2f}"))

# ---------------------------------------------------------------- 2019 vs 2025 (annual weighted)
def annual(year):
    g = df[df["year"] == year]
    return dict(
        w_min_nom=wmean(g, "pmin"), w_avg_nom=wmean(g, "pavg"), w_max_nom=wmean(g, "pmax"),
        n=len(g))
def deflate_year(year, val):
    # deflate a nominal annual figure by that year's avg CPI
    qs = [f"{year}Q{i}" for i in range(1,5)]
    cpi_yr = np.mean([cpi_q[q] for q in qs])
    return val * cpi_2025_avg / cpi_yr

a19, a25 = annual(2019), annual(2025)
for a, yr in [(a19,2019),(a25,2025)]:
    a["w_min_real"] = deflate_year(yr, a["w_min_nom"])
    a["w_avg_real"] = deflate_year(yr, a["w_avg_nom"])
    a["w_max_real"] = deflate_year(yr, a["w_max_nom"])
    a["spread"] = a["w_max_nom"]/a["w_min_nom"]

print("\n=== Annual weighted, 2019 vs 2025 ===")
for lab,a in [("2019",a19),("2025",a25)]:
    print(f"{lab}: min_real={a['w_min_real']:.2f} avg_real={a['w_avg_real']:.2f} "
          f"max_real={a['w_max_real']:.2f} spread={a['spread']:.3f} (n={a['n']:,})")

def pct(new, old): return 100*(new/old-1)
summary = {
 "min_real_2019": a19["w_min_real"], "min_real_2025": a25["w_min_real"],
 "avg_real_2019": a19["w_avg_real"], "avg_real_2025": a25["w_avg_real"],
 "max_real_2019": a19["w_max_real"], "max_real_2025": a25["w_max_real"],
 "min_nom_2019": a19["w_min_nom"], "min_nom_2025": a25["w_min_nom"],
 "max_nom_2019": a19["w_max_nom"], "max_nom_2025": a25["w_max_nom"],
 "spread_2019": a19["spread"], "spread_2025": a25["spread"],
 "d_min_real_pct": pct(a25["w_min_real"], a19["w_min_real"]),
 "d_avg_real_pct": pct(a25["w_avg_real"], a19["w_avg_real"]),
 "d_max_real_pct": pct(a25["w_max_real"], a19["w_max_real"]),
 "d_min_nom_pct": pct(a25["w_min_nom"], a19["w_min_nom"]),
 "d_max_nom_pct": pct(a25["w_max_nom"], a19["w_max_nom"]),
 "d_spread_pct": pct(a25["spread"], a19["spread"]),
 "spread_pp": a25["spread"]-a19["spread"],
}
print("\n=== Change 2019 -> 2025 ===")
print(f"get-in (min) real: {summary['d_min_real_pct']:+.1f}%   nominal: {summary['d_min_nom_pct']:+.1f}%")
print(f"avg price   real: {summary['d_avg_real_pct']:+.1f}%")
print(f"top (max)   real: {summary['d_max_real_pct']:+.1f}%   nominal: {summary['d_max_nom_pct']:+.1f}%")
print(f"spread ratio: {summary['spread_2019']:.2f}x -> {summary['spread_2025']:.2f}x "
      f"({summary['d_spread_pct']:+.1f}%, {summary['spread_pp']:+.2f}x)")

# ---------------------------------------------------------------- charts
QOK = res["quarter"].tolist()
x = np.arange(len(QOK))
thin_mask = res["thin"].values

def qlabels():
    return [q if q.endswith("Q1") else "" for q in QOK]

def draw_indexed(col_suffix, title, fname):
    fig, ax = plt.subplots(figsize=(11,6))
    styles = {"min":("Get-in (min)","#1f77b4"),
              "avg":("Average","#2ca02c"),
              "max":("Top (max)","#d62728")}
    for s,(lab,c) in styles.items():
        ax.plot(x, res[f"idx_{s}_{col_suffix}"], marker="o", ms=3, color=c, label=lab)
    ax.axhline(100, color="grey", lw=.8, ls="--")
    for xi in x[thin_mask]:
        ax.axvspan(xi-0.5, xi+0.5, color="orange", alpha=0.12, zorder=0)
    ax.set_xticks(x); ax.set_xticklabels(qlabels(), rotation=45, ha="right", fontsize=8)
    ax.set_ylabel("Index (2019Q1 = 100)")
    ax.set_title(title)
    ax.legend(frameon=False); ax.grid(alpha=.25)
    ax.text(0.01,0.02,"shaded = thin sample (COVID, n<2000)",transform=ax.transAxes,
            fontsize=7,color="darkorange")
    fig.tight_layout(); fig.savefig(fname, dpi=130); plt.close(fig)

draw_indexed("real", "US ticket prices, attendance-weighted, real 2025$ (indexed 2019Q1=100)",
             f"{DIR}/qtr_indexed_real.png")
draw_indexed("nom",  "US ticket prices, attendance-weighted, NOMINAL (indexed 2019Q1=100)",
             f"{DIR}/qtr_indexed_nominal.png")

# spread ratio standalone
fig, ax = plt.subplots(figsize=(11,6))
ax.plot(x, res["spread_real"], marker="o", ms=3, color="#6a3d9a", label="Max / Min spread ratio")
for xi in x[thin_mask]:
    ax.axvspan(xi-0.5, xi+0.5, color="orange", alpha=0.12, zorder=0)
# trend line over non-thin quarters
ok = ~thin_mask
z = np.polyfit(x[ok], res["spread_real"].values[ok], 1)
ax.plot(x, np.polyval(z, x), color="grey", ls="--", lw=1,
        label=f"trend (+{z[0]:.3f}x/qtr)")
ax.set_xticks(x); ax.set_xticklabels(qlabels(), rotation=45, ha="right", fontsize=8)
ax.set_ylabel("Spread ratio  (weighted max ÷ weighted min)")
ax.set_title("US ticket-price spread ratio (max/min), attendance-weighted")
ax.legend(frameon=False); ax.grid(alpha=.25)
ax.text(0.01,0.02,"unit-free: identical in nominal & real terms",transform=ax.transAxes,fontsize=7,color="grey")
fig.tight_layout(); fig.savefig(f"{DIR}/qtr_spread_ratio.png", dpi=130); plt.close(fig)
print("\ncharts written")

# ---------------------------------------------------------------- Excel
wb = Workbook()
HDR = Font(bold=True, color="FFFFFF"); HFILL = PatternFill("solid", fgColor="305496")
THIN = PatternFill("solid", fgColor="FCE4D6")
TITLE = Font(bold=True, size=13)
thin_border = Border(*(Side(style="thin", color="D9D9D9"),)*4)

def style_header(ws, ncol, r=1):
    for c in range(1, ncol+1):
        cell = ws.cell(r, c); cell.font = HDR; cell.fill = HFILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def autosize(ws, widths):
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w

# --- Summary sheet
ws = wb.active; ws.title = "Summary"
ws["A1"] = "US Ticket Prices — Quarterly, Attendance-Weighted (2019+)"; ws["A1"].font = TITLE
notes = [
 "",
 "Scope: country = United States, eventDate 2019Q1–2025Q4. Forward-dated 2026+ shows excluded.",
 f"Events in panel: {len(df):,}.  Weighting: attendance (ticketsSold) — larger shows count more.",
 "Prices already USD (Pollstar ticketPriceMin/Avg/Max). Winsorized each field to [p0.1, p99.9] to remove data-entry errors.",
 "Deflator: CPI-U (CPIAUCSL, seasonally adjusted, FRED). Base = 2025 annual average. 'Real' = constant 2025 USD.",
 "Spread ratio = weighted max ÷ weighted min (unit-free; identical nominal vs real).",
 "Thin quarters (n_events < 2000) are the COVID shutdown: 2020Q2–2021Q2. Flagged/shaded throughout.",
 "",
 "HEADLINE — change 2019 → 2025 (full-year weighted):",
]
r = 3
for n in notes:
    ws.cell(r,1,n); r+=1
tbl = [
 ["Metric","2019","2025","Δ real","Δ nominal"],
 ["Get-in / min price (real 2025$)", f"${summary['min_real_2019']:.2f}", f"${summary['min_real_2025']:.2f}",
    f"{summary['d_min_real_pct']:+.1f}%", f"{summary['d_min_nom_pct']:+.1f}%"],
 ["Average price (real 2025$)", f"${summary['avg_real_2019']:.2f}", f"${summary['avg_real_2025']:.2f}",
    f"{summary['d_avg_real_pct']:+.1f}%", "—"],
 ["Top / max price (real 2025$)", f"${summary['max_real_2019']:.2f}", f"${summary['max_real_2025']:.2f}",
    f"{summary['d_max_real_pct']:+.1f}%", f"{summary['d_max_nom_pct']:+.1f}%"],
 ["Spread ratio (max/min)", f"{summary['spread_2019']:.2f}x", f"{summary['spread_2025']:.2f}x",
    f"{summary['d_spread_pct']:+.1f}%  ({summary['spread_pp']:+.2f}x)", "(same)"],
]
r += 1; top = r
for row in tbl:
    for c,val in enumerate(row,1):
        ws.cell(r,c,val)
    r += 1
style_header(ws, 5, top)
verdict = ("WIDENING" if summary['spread_pp']>0 else "NARROWING")
ws.cell(r+1,1, f"Verdict: the max/min spread is {verdict} — "
              f"{summary['spread_2019']:.2f}x (2019) → {summary['spread_2025']:.2f}x (2025), "
              f"{summary['d_spread_pct']:+.1f}%.").font = Font(bold=True)
autosize(ws, [42,12,12,20,14])

# --- data sheets
def write_table(ws, cols, headers, widths, numfmt):
    ws.append(headers); style_header(ws, len(headers))
    for _,rr in res.iterrows():
        vals = [rr[c] for c in cols]
        ws.append(vals)
    # formatting
    for i in range(2, len(res)+2):
        if res.iloc[i-2]["thin"]:
            for c in range(1, len(headers)+1):
                ws.cell(i,c).fill = THIN
        for c,fmt in numfmt.items():
            ws.cell(i,c).number_format = fmt
    autosize(ws, widths)
    ws.freeze_panes = "A2"

ws2 = wb.create_sheet("Quarterly_Real_2025USD")
write_table(ws2,
    ["quarter","n_events","tickets","w_min_real","w_avg_real","w_max_real","spread_real",
     "idx_min_real","idx_avg_real","idx_max_real","thin"],
    ["Quarter","Events","Tickets","Min (get-in)","Avg","Max","Spread max/min",
     "Idx Min","Idx Avg","Idx Max","Thin?"],
    [10,9,12,14,10,10,15,10,10,10,7],
    {3:"#,##0",4:'"$"#,##0.00',5:'"$"#,##0.00',6:'"$"#,##0.00',7:"0.00",8:"0.0",9:"0.0",10:"0.0"})

ws3 = wb.create_sheet("Quarterly_Nominal")
write_table(ws3,
    ["quarter","n_events","tickets","w_min_nom","w_avg_nom","w_max_nom","spread_nom",
     "idx_min_nom","idx_avg_nom","idx_max_nom","thin"],
    ["Quarter","Events","Tickets","Min (get-in)","Avg","Max","Spread max/min",
     "Idx Min","Idx Avg","Idx Max","Thin?"],
    [10,9,12,14,10,10,15,10,10,10,7],
    {3:"#,##0",4:'"$"#,##0.00',5:'"$"#,##0.00',6:'"$"#,##0.00',7:"0.00",8:"0.0",9:"0.0",10:"0.0"})

# --- charts sheet
wsC = wb.create_sheet("Charts")
wsC["A1"]="Indexed price lines (2019Q1=100) & spread ratio"; wsC["A1"].font=TITLE
img_specs = [("qtr_indexed_real.png","A3"),
             ("qtr_indexed_nominal.png","A34"),
             ("qtr_spread_ratio.png","A65")]
for fn,anchor in img_specs:
    img = XLImage(f"{DIR}/{fn}"); img.anchor=anchor; wsC.add_image(img)

out = f"{DIR}/US_Quarterly_Ticket_Price_Spread.xlsx"
wb.save(out)
print("saved", out)
