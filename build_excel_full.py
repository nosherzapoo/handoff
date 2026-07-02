#!/usr/bin/env python3
"""
Master build for US_Quarterly_Ticket_Price_Spread.xlsx.

Analysis 1  (P1): quarterly attendance-weighted min/avg/max prices + max/min spread,
                  CPI-deflated to 2025$, nominal & real, indexed to 2019Q1=100.
Analysis 2  (P2): within-year gross tiers (Top1% / Top2-5% / Mid 6-50% / Bottom 50%):
                  gross share, ticket share, avg real ticketPriceAvg, avg capacity sold.
Analysis 3  (P3): revenue concentration by year — Gini/HHI of grossUSD across events
                  and across headliners, plus top-10 headliner share of gross.

US only, eventDate 2019Q1-2025Q4 (forward-dated 2026+ shows excluded).
Prices already USD. Deflator: CPI-U CPIAUCSL (SA, FRED), base = 2025 annual avg.
"""
import csv
from datetime import datetime
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DIR = "/Users/nosherzapoo/Desktop/claude/pollstar"

# ================================================================ CPI-U
cpi_m = {}
with open(f"{DIR}/cpi_u_cpiaucsl.csv") as f:
    for row in csv.DictReader(f):
        v = row["CPIAUCSL"].strip()
        if v:
            cpi_m[row["observation_date"][:7]] = float(v)
if "2025-10" not in cpi_m:                       # gov-shutdown gap -> interpolate
    cpi_m["2025-10"] = (cpi_m["2025-09"] + cpi_m["2025-11"]) / 2

cpi_q, tmp = {}, {}
for ym, v in cpi_m.items():
    y, m = int(ym[:4]), int(ym[5:7])
    tmp.setdefault((y, (m - 1) // 3 + 1), []).append(v)
for (y, q), vals in tmp.items():
    cpi_q[f"{y}Q{q}"] = sum(vals) / len(vals)
cpi_yr = {y: np.mean([cpi_m[f"{y}-{mm:02d}"] for mm in range(1, 13)]) for y in range(2019, 2026)}
CPI_BASE = cpi_yr[2025]                           # constant 2025 USD
print(f"CPI-U 2025 annual avg (base) = {CPI_BASE:.3f}")

# ================================================================ load master data
recs = []
with open(f"{DIR}/concerts.csv") as f:
    for row in csv.DictReader(f):
        if row["country"] != "United States":
            continue
        try:
            d = datetime.strptime(row["eventDate"], "%m/%d/%Y")
        except ValueError:
            continue
        if d.year < 2019 or d.year >= 2026:
            continue
        def num(k):
            try:
                return float(row[k])
            except ValueError:
                return np.nan
        recs.append({
            "year": d.year, "quarter": f"{d.year}Q{(d.month-1)//3+1}",
            "ts": num("ticketsSold"), "pmin": num("ticketPriceMin"),
            "pavg": num("ticketPriceAvg"), "pmax": num("ticketPriceMax"),
            "gross": num("grossUSD"), "cap": num("avgCapacitySold"),
            "head": row["headLiner"].strip(),
        })
M = pd.DataFrame(recs)
M["defl_yr"] = M["year"].map(lambda y: CPI_BASE / cpi_yr[y])   # nominal -> real(2025)
M["defl_q"]  = M["quarter"].map(lambda q: CPI_BASE / cpi_q[q])
print(f"US master rows 2019-2025: {len(M):,}")

def gini(x):
    """Gini coefficient of a 1-D array of non-negative values."""
    x = np.sort(np.asarray(x, float))
    n = len(x)
    if n == 0 or x.sum() == 0:
        return np.nan
    idx = np.arange(1, n + 1)
    return (2 * np.sum(idx * x) / (n * x.sum())) - (n + 1) / n

# ================================================================ P1: quarterly weighted prices
p1 = M.dropna(subset=["ts", "pmin", "pavg", "pmax"]).copy()
p1 = p1[(p1.ts > 0) & (p1.pmin > 0) & (p1.pavg > 0) & (p1.pmax > 0) & (p1.pmax >= p1.pmin)]
for col in ["pmin", "pavg", "pmax"]:            # kill data-entry outliers
    lo, hi = p1[col].quantile([0.001, 0.999])
    p1[col] = p1[col].clip(lo, hi)
print(f"P1 valid price events: {len(p1):,}")

def wmean(g, c): return np.average(g[c], weights=g["ts"])
rows = []
for q, g in p1.groupby("quarter"):
    rows.append({"quarter": q, "n_events": len(g), "tickets": g.ts.sum(),
                 "w_min_nom": wmean(g, "pmin"), "w_avg_nom": wmean(g, "pavg"),
                 "w_max_nom": wmean(g, "pmax")})
res = pd.DataFrame(rows).sort_values("quarter").reset_index(drop=True)
res["defl"] = res["quarter"].map(lambda q: CPI_BASE / cpi_q[q])
for s in ["min", "avg", "max"]:
    res[f"w_{s}_real"] = res[f"w_{s}_nom"] * res["defl"]
res["spread_nom"]  = res.w_max_nom / res.w_min_nom
res["spread_real"] = res.w_max_real / res.w_min_real
res["thin"] = res.n_events < 2000
base = res[res.quarter == "2019Q1"].iloc[0]
for s in ["min", "avg", "max"]:
    res[f"idx_{s}_real"] = 100 * res[f"w_{s}_real"] / base[f"w_{s}_real"]
    res[f"idx_{s}_nom"]  = 100 * res[f"w_{s}_nom"]  / base[f"w_{s}_nom"]

def wyr(year, c):
    g = p1[p1.year == year]; return np.average(g[c], weights=g["ts"])
a = {}
for yr in (2019, 2025):
    a[yr] = {"min": wyr(yr, "pmin") * CPI_BASE / cpi_yr[yr],
             "avg": wyr(yr, "pavg") * CPI_BASE / cpi_yr[yr],
             "max": wyr(yr, "pmax") * CPI_BASE / cpi_yr[yr],
             "min_nom": wyr(yr, "pmin"), "max_nom": wyr(yr, "pmax"),
             "spread": wyr(yr, "pmax") / wyr(yr, "pmin")}
pct = lambda n, o: 100 * (n / o - 1)
P1 = {
 "min19": a[2019]["min"], "min25": a[2025]["min"], "avg19": a[2019]["avg"], "avg25": a[2025]["avg"],
 "max19": a[2019]["max"], "max25": a[2025]["max"], "sp19": a[2019]["spread"], "sp25": a[2025]["spread"],
 "dmin_r": pct(a[2025]["min"], a[2019]["min"]), "dmin_n": pct(a[2025]["min_nom"], a[2019]["min_nom"]),
 "davg_r": pct(a[2025]["avg"], a[2019]["avg"]),
 "dmax_r": pct(a[2025]["max"], a[2019]["max"]), "dmax_n": pct(a[2025]["max_nom"], a[2019]["max_nom"]),
 "dsp": pct(a[2025]["spread"], a[2019]["spread"]), "sp_pp": a[2025]["spread"] - a[2019]["spread"],
}

# ================================================================ P2: within-year gross tiers
p2 = M.dropna(subset=["gross"]).copy()
p2 = p2[p2.gross > 0]
TIERS = ["Top 1%", "Top 2-5%", "Mid (6-50%)", "Bottom 50%"]

def assign_tier(g):
    q99, q95, q50 = g.quantile([0.99, 0.95, 0.50])
    t = np.where(g >= q99, "Top 1%",
        np.where(g >= q95, "Top 2-5%",
        np.where(g >= q50, "Mid (6-50%)", "Bottom 50%")))
    return t
p2["tier"] = pd.Series(index=p2.index, dtype="object")
for yr, g in p2.groupby("year"):
    p2.loc[g.index, "tier"] = assign_tier(g["gross"])
p2["cap_c"] = p2["cap"].clip(0, 100)             # cap capacity% at 100 (data errors up to 816)

tier_rows = []
for yr in range(2019, 2026):
    gy = p2[p2.year == yr]
    tot_g, tot_t = gy.gross.sum(), gy.ts.sum(skipna=True)
    for t in TIERS:
        gt = gy[gy.tier == t]
        pav = gt.pavg.dropna(); pav = pav[pav > 0]
        tier_rows.append({
            "year": yr, "tier": t, "n_events": len(gt),
            "share_gross": gt.gross.sum() / tot_g,
            "share_tickets": gt.ts.sum(skipna=True) / tot_t,
            "avg_pavg_real": (pav.mean() * CPI_BASE / cpi_yr[yr]) if len(pav) else np.nan,
            "avg_cap": gt.cap_c.mean(),
        })
T = pd.DataFrame(tier_rows)
# cumulative Top-5% = Top1% + Top2-5%
def share_for(yr, tiers):
    return T[(T.year == yr) & (T.tier.isin(tiers))]["share_gross"].sum()
P2 = {"y": list(range(2019, 2026))}
P2["top1"]   = [share_for(y, ["Top 1%"]) for y in P2["y"]]
P2["top5"]   = [share_for(y, ["Top 1%", "Top 2-5%"]) for y in P2["y"]]
P2["mid"]    = [share_for(y, ["Mid (6-50%)"]) for y in P2["y"]]
P2["bottom"] = [share_for(y, ["Bottom 50%"]) for y in P2["y"]]

# ================================================================ P3: concentration
p3 = M[M.gross > 0].copy()
conc_rows = []
for yr in range(2019, 2026):
    gy = p3[p3.year == yr]
    ev = gy.gross.values
    by_head = gy.groupby("head")["gross"].sum()
    tot = by_head.sum()
    top10 = by_head.sort_values(ascending=False).head(10).sum() / tot
    hhi_head = float(((by_head / tot) ** 2).sum())   # 0..1 (Herfindahl)
    conc_rows.append({
        "year": yr, "n_events": len(gy), "n_headliners": by_head.size,
        "gini_events": gini(ev), "gini_headliners": gini(by_head.values),
        "hhi_headliners": hhi_head, "top10_head_share": top10,
        "total_gross": tot,
    })
C = pd.DataFrame(conc_rows)
P3 = {
 "gini_ev_19": C.loc[C.year == 2019, "gini_events"].iat[0],
 "gini_ev_25": C.loc[C.year == 2025, "gini_events"].iat[0],
 "gini_h_19": C.loc[C.year == 2019, "gini_headliners"].iat[0],
 "gini_h_25": C.loc[C.year == 2025, "gini_headliners"].iat[0],
 "top10_19": C.loc[C.year == 2019, "top10_head_share"].iat[0],
 "top10_25": C.loc[C.year == 2025, "top10_head_share"].iat[0],
}

# ================================================================ P4: top-1% ARTISTS get-in vs max, quarterly
p4 = M[M.gross > 0].copy()
top_sets = {}                                    # top 1% of headliners per year, by annual gross
for yr in range(2019, 2026):
    hs = p4[p4.year == yr].groupby("head")["gross"].sum()
    top_sets[yr] = set(hs[hs >= hs.quantile(0.99)].index)
p4["is_top1_artist"] = False
for yr, s in top_sets.items():
    m = p4.year == yr
    p4.loc[m, "is_top1_artist"] = p4.loc[m, "head"].isin(s)

q4 = p4[p4.is_top1_artist].dropna(subset=["ts", "pmin", "pmax"]).copy()
q4 = q4[(q4.ts > 0) & (q4.pmin > 0) & (q4.pmax > 0) & (q4.pmax >= q4.pmin)]
for col in ["pmin", "pmax"]:                     # kill data-entry outliers within this subset
    lo, hi = q4[col].quantile([0.001, 0.999]); q4[col] = q4[col].clip(lo, hi)

rows = []
for q, g in q4.groupby("quarter"):
    rows.append({"quarter": q, "n_events": len(g), "n_artists": g["head"].nunique(), "tickets": g.ts.sum(),
                 "w_min": np.average(g.pmin, weights=g.ts), "w_max": np.average(g.pmax, weights=g.ts),
                 "s_min": g.pmin.mean(), "s_max": g.pmax.mean()})
Q4 = pd.DataFrame(rows).sort_values("quarter").reset_index(drop=True)
Q4["year"] = Q4.quarter.str[:4].astype(int)
Q4["defl"] = Q4.quarter.map(lambda q: CPI_BASE / cpi_q[q])
Q4["w_min_real"] = Q4.w_min * Q4.defl; Q4["w_max_real"] = Q4.w_max * Q4.defl
Q4["ratio_w"] = Q4.w_max / Q4.w_min             # attendance-weighted max/min (unit-free)
Q4["ratio_s"] = Q4.s_max / Q4.s_min             # simple-average max/min
Q4["thin"] = Q4.n_events < 50                    # top-artist events are sparse; COVID quarters collapse

def a_ratio(yr):                                 # annual attendance-weighted
    g = q4[q4.quarter.str.startswith(str(yr))]
    wmn = np.average(g.pmin, weights=g.ts); wmx = np.average(g.pmax, weights=g.ts)
    return wmn * CPI_BASE / cpi_yr[yr], wmx * CPI_BASE / cpi_yr[yr], wmx / wmn
P4 = {}
P4["min19"], P4["max19"], P4["r19"] = a_ratio(2019)
P4["min25"], P4["max25"], P4["r25"] = a_ratio(2025)
ok4 = ~Q4.thin.values
P4["slope"] = np.polyfit(np.arange(len(Q4))[ok4], Q4.ratio_w.values[ok4], 1)[0]

# ================================================================ console report
print("\n=== P4 top-1% ARTISTS: get-in vs max, quarterly (real 2025$) ===")
print(Q4[["quarter","n_events","n_artists","w_min_real","w_max_real","ratio_w","ratio_s","thin"]]
      .to_string(index=False, float_format=lambda x: f"{x:,.2f}"))
print(f"Annual weighted ratio: 2019={P4['r19']:.2f}x  2025={P4['r25']:.2f}x  "
      f"(Δ {(P4['r25']/P4['r19']-1)*100:+.1f}%); trend {P4['slope']:+.4f}x/qtr")
print("\n=== P2 gross-share by tier (2019 vs 2025) ===")
for lab, key in [("Top 1%", "top1"), ("Top 5% (cum)", "top5"), ("Mid 6-50%", "mid"), ("Bottom 50%", "bottom")]:
    v = P2[key]
    print(f"{lab:14s} 2019={v[0]*100:5.1f}%  2025={v[-1]*100:5.1f}%  Δ={(v[-1]-v[0])*100:+.1f}pp")
print("\n=== P3 concentration ===")
print(C[["year","gini_events","gini_headliners","hhi_headliners","top10_head_share"]]
      .to_string(index=False, float_format=lambda x: f"{x:,.4f}"))
print(f"\nHeadliner Gini {P3['gini_h_19']:.3f} -> {P3['gini_h_25']:.3f}; "
      f"Top-10 headliner gross share {P3['top10_19']*100:.1f}% -> {P3['top10_25']*100:.1f}%")

# ================================================================ CHARTS
QOK = res.quarter.tolist(); x = np.arange(len(QOK)); thin = res.thin.values
qlab = [q if q.endswith("Q1") else "" for q in QOK]

def draw_indexed(sfx, title, fname):
    fig, ax = plt.subplots(figsize=(11, 6))
    for s, (lab, c) in {"min": ("Get-in (min)", "#1f77b4"), "avg": ("Average", "#2ca02c"),
                        "max": ("Top (max)", "#d62728")}.items():
        ax.plot(x, res[f"idx_{s}_{sfx}"], marker="o", ms=3, color=c, label=lab)
    ax.axhline(100, color="grey", lw=.8, ls="--")
    for xi in x[thin]:
        ax.axvspan(xi - .5, xi + .5, color="orange", alpha=.12, zorder=0)
    ax.set_xticks(x); ax.set_xticklabels(qlab, rotation=45, ha="right", fontsize=8)
    ax.set_ylabel("Index (2019Q1 = 100)"); ax.set_title(title)
    ax.legend(frameon=False); ax.grid(alpha=.25)
    ax.text(.01, .02, "shaded = thin sample (COVID, n<2000)", transform=ax.transAxes,
            fontsize=7, color="darkorange")
    fig.tight_layout(); fig.savefig(fname, dpi=130); plt.close(fig)

draw_indexed("real", "US ticket prices, attendance-weighted, real 2025$ (indexed 2019Q1=100)",
             f"{DIR}/qtr_indexed_real.png")
draw_indexed("nom", "US ticket prices, attendance-weighted, NOMINAL (indexed 2019Q1=100)",
             f"{DIR}/qtr_indexed_nominal.png")

fig, ax = plt.subplots(figsize=(11, 6))
ax.plot(x, res.spread_real, marker="o", ms=3, color="#6a3d9a", label="Max / Min spread ratio")
for xi in x[thin]:
    ax.axvspan(xi - .5, xi + .5, color="orange", alpha=.12, zorder=0)
ok = ~thin; z = np.polyfit(x[ok], res.spread_real.values[ok], 1)
ax.plot(x, np.polyval(z, x), color="grey", ls="--", lw=1, label=f"trend ({z[0]:+.3f}x/qtr)")
ax.set_xticks(x); ax.set_xticklabels(qlab, rotation=45, ha="right", fontsize=8)
ax.set_ylabel("Spread ratio (weighted max ÷ weighted min)")
ax.set_title("US ticket-price spread ratio (max/min), attendance-weighted")
ax.legend(frameon=False); ax.grid(alpha=.25)
ax.text(.01, .02, "unit-free: identical in nominal & real terms", transform=ax.transAxes, fontsize=7, color="grey")
fig.tight_layout(); fig.savefig(f"{DIR}/qtr_spread_ratio.png", dpi=130); plt.close(fig)

# --- P2 stacked gross-share chart
yrs = P2["y"]; xb = np.arange(len(yrs))
fig, ax = plt.subplots(figsize=(10, 6))
bottoms = np.zeros(len(yrs))
stack = [("Bottom 50%", P2["bottom"], "#c6dbef"),
         ("Mid (6-50%)", P2["mid"], "#6baed6"),
         ("Top 2-5%", [P2["top5"][i] - P2["top1"][i] for i in range(len(yrs))], "#fd8d3c"),
         ("Top 1%", P2["top1"], "#a50f15")]
for lab, vals, c in stack:
    vals = np.array(vals)
    ax.bar(xb, vals * 100, bottom=bottoms * 100, color=c, label=lab, edgecolor="white", width=.72)
    for i, v in enumerate(vals):
        if v > 0.03:
            ax.text(xb[i], (bottoms[i] + v / 2) * 100, f"{v*100:.0f}%",
                    ha="center", va="center", fontsize=8, color="white" if c in ("#a50f15","#6baed6") else "black")
    bottoms += vals
ax.set_xticks(xb); ax.set_xticklabels(yrs)
ax.set_ylabel("Share of total US gross (%)"); ax.set_ylim(0, 100)
ax.set_title("US concert gross by within-year event tier (share of total gross)")
ax.legend(frameon=False, ncol=4, loc="lower center", bbox_to_anchor=(.5, -.13))
fig.tight_layout(); fig.savefig(f"{DIR}/tier_stacked_share.png", dpi=130); plt.close(fig)

# --- P3 concentration charts
fig, (axL, axR) = plt.subplots(1, 2, figsize=(12, 5))
axL.plot(C.year, C.gini_headliners, marker="o", color="#a50f15", label="Gini across headliners")
axL.plot(C.year, C.gini_events, marker="s", color="#4575b4", label="Gini across events")
axL.set_ylabel("Gini coefficient"); axL.set_title("Revenue concentration (Gini)")
axL.set_xticks(list(yrs)); axL.grid(alpha=.25); axL.legend(frameon=False)
axR.plot(C.year, C.top10_head_share * 100, marker="o", color="#238b45")
for xi, yv in zip(C.year, C.top10_head_share * 100):
    axR.annotate(f"{yv:.1f}%", (xi, yv), textcoords="offset points", xytext=(0, 6), fontsize=8, ha="center")
axR.set_ylabel("% of total US gross"); axR.set_title("Top-10 headliners' share of US gross")
axR.set_xticks(list(yrs)); axR.grid(alpha=.25)
fig.tight_layout(); fig.savefig(f"{DIR}/concentration.png", dpi=130); plt.close(fig)

# --- P4 top-1% artists: get-in vs max levels + ratio
xq = np.arange(len(Q4)); thin4 = Q4.thin.values
q4lab = [q if q.endswith("Q1") else "" for q in Q4.quarter]
fig, axL = plt.subplots(figsize=(11, 6))
axL.plot(xq, Q4.w_max_real, marker="o", ms=3, color="#d62728", label="Top price (max), real 2025$")
axL.plot(xq, Q4.w_min_real, marker="o", ms=3, color="#1f77b4", label="Get-in (min), real 2025$")
axL.set_ylabel("Attendance-weighted price (real 2025$)")
axR = axL.twinx()
axR.plot(xq, Q4.ratio_w, marker="s", ms=3, color="#6a3d9a", lw=2, label="Max/Min ratio (right axis)")
zz = np.polyfit(xq[~thin4], Q4.ratio_w.values[~thin4], 1)
axR.plot(xq, np.polyval(zz, xq), color="#6a3d9a", ls="--", lw=1, alpha=.7)
axR.set_ylabel("Max / Min ratio", color="#6a3d9a"); axR.tick_params(axis="y", labelcolor="#6a3d9a")
for xi in xq[thin4]:
    axL.axvspan(xi - .5, xi + .5, color="orange", alpha=.12, zorder=0)
axL.set_xticks(xq); axL.set_xticklabels(q4lab, rotation=45, ha="right", fontsize=8)
axL.set_title("Top 1% of artists (by annual gross): get-in vs top price, attendance-weighted")
l1, la1 = axL.get_legend_handles_labels(); l2, la2 = axR.get_legend_handles_labels()
axL.legend(l1 + l2, la1 + la2, frameon=False, loc="upper left", fontsize=8)
axL.grid(alpha=.25)
axL.text(.01, .02, "shaded = thin sample (n<50 events, COVID)", transform=axL.transAxes, fontsize=7, color="darkorange")
fig.tight_layout(); fig.savefig(f"{DIR}/top1_artists_ratio.png", dpi=130); plt.close(fig)
print("charts written")

# ================================================================ EXCEL
wb = Workbook()
HDR = Font(bold=True, color="FFFFFF"); HFILL = PatternFill("solid", fgColor="305496")
THIN_F = PatternFill("solid", fgColor="FCE4D6"); TITLE = Font(bold=True, size=13)
SUB = Font(bold=True, size=11); WRAP = Alignment(wrap_text=True, vertical="top")

def hdr(ws, ncol, r=1):
    for c in range(1, ncol + 1):
        cell = ws.cell(r, c); cell.font = HDR; cell.fill = HFILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def widths(ws, w):
    for i, wv in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = wv

# ---- README
ws = wb.active; ws.title = "README"
ws["A1"] = "US Live-Music Ticket-Price & Gross Concentration — Methodology"; ws["A1"].font = TITLE
lines = [
 "",
 "SCOPE (all analyses): country = United States; eventDate 2019Q1–2025Q4. Forward-dated 2026+ shows excluded.",
 "Prices (ticketPriceMin/Avg/Max) are already USD. Real = constant 2025 USD via CPI-U (CPIAUCSL, seasonally adjusted, FRED);",
 f"   base = 2025 annual average CPI ({CPI_BASE:.1f}). Oct-2025 CPI interpolated (FRED gap).",
 "",
 "ANALYSIS 1 — Quarterly weighted prices & max/min spread  [sheets P1_*]",
 f"   • Attendance-weighted (weight = ticketsSold) mean of get-in / avg / top price per quarter. n={len(p1):,} valid-price events.",
 "   • Each price field winsorized to [p0.1, p99.9] to remove data-entry errors. Spread ratio = weighted max ÷ weighted min (unit-free).",
 "   • Indexed to 100 at 2019Q1; shown nominal & real. Thin quarters (n<2000 = COVID 2020Q2–2021Q2) flagged.",
 "",
 "ANALYSIS 2 — Within-year gross tiers  [sheets P2_*]",
 f"   • Events ranked by grossUSD within each year and split into Top 1% / Top 2–5% / Mid (6–50%) / Bottom 50%. n={len(p2):,} events.",
 "   • Per tier×year: share of total gross, share of total tickets, avg ticketPriceAvg (real 2025$), avg avgCapacitySold (capped 100%).",
 "   • 'Top 5%' in headline = cumulative Top 1% + Top 2–5%.",
 "",
 "ANALYSIS 3 — Revenue concentration  [sheets P3_*]",
 f"   • Per year: Gini & Herfindahl of grossUSD across events and across headliners (gross summed by headLiner). n={len(p3):,} events.",
 "   • Gini 0 = perfectly equal, 1 = one event/artist takes all. Plus top-10 headliners' share of total annual gross.",
 "",
 "ANALYSIS 4 — Top 1% of artists: get-in vs top price  [sheets P4_*]",
 "   • Top 1% of headliners by gross (ranked within each year). Their events' attendance-weighted get-in (min) & top (max) price per quarter,",
 "   • real 2025$, plus max/min ratio to test whether the premium-vs-cheap-seat gap is widening for the biggest acts. Thin quarters (n<50) flagged.",
]
r = 3
for t in lines:
    c = ws.cell(r, 1, t)
    if t.startswith("ANALYSIS"): c.font = SUB
    r += 1
widths(ws, [120])

# ---- P1_Summary
ws = wb.create_sheet("P1_Summary")
ws["A1"] = "Analysis 1 — Quarterly attendance-weighted prices & spread"; ws["A1"].font = TITLE
ws["A3"] = "Change 2019 → 2025 (full-year attendance-weighted):"; ws["A3"].font = SUB
tbl = [["Metric", "2019", "2025", "Δ real", "Δ nominal"],
 ["Get-in / min (real 2025$)", f"${P1['min19']:.2f}", f"${P1['min25']:.2f}", f"{P1['dmin_r']:+.1f}%", f"{P1['dmin_n']:+.1f}%"],
 ["Average (real 2025$)", f"${P1['avg19']:.2f}", f"${P1['avg25']:.2f}", f"{P1['davg_r']:+.1f}%", "—"],
 ["Top / max (real 2025$)", f"${P1['max19']:.2f}", f"${P1['max25']:.2f}", f"{P1['dmax_r']:+.1f}%", f"{P1['dmax_n']:+.1f}%"],
 ["Spread ratio (max/min)", f"{P1['sp19']:.2f}x", f"{P1['sp25']:.2f}x", f"{P1['dsp']:+.1f}% ({P1['sp_pp']:+.2f}x)", "(same)"]]
r0 = 5
for i, row in enumerate(tbl):
    for c, v in enumerate(row, 1):
        ws.cell(r0 + i, c, v)
hdr(ws, 5, r0)
ws.cell(r0 + len(tbl) + 1, 1,
        f"Verdict: spread is essentially FLAT (not widening): {P1['sp19']:.2f}x → {P1['sp25']:.2f}x ({P1['dsp']:+.1f}%). "
        "Pattern is seasonal (Q3 peaks, Q1 troughs), not trending.").font = SUB
widths(ws, [30, 12, 12, 20, 14])

# ---- P1 quarterly tables
def write_res(ws, cols, heads, w, fmt):
    ws.append(heads); hdr(ws, len(heads))
    for _, rr in res.iterrows():
        ws.append([rr[c] for c in cols])
    for i in range(2, len(res) + 2):
        if res.iloc[i - 2].thin:
            for c in range(1, len(heads) + 1):
                ws.cell(i, c).fill = THIN_F
        for c, f in fmt.items():
            ws.cell(i, c).number_format = f
    widths(ws, w); ws.freeze_panes = "A2"

cols_r = ["quarter","n_events","tickets","w_min_real","w_avg_real","w_max_real","spread_real","idx_min_real","idx_avg_real","idx_max_real","thin"]
cols_n = ["quarter","n_events","tickets","w_min_nom","w_avg_nom","w_max_nom","spread_nom","idx_min_nom","idx_avg_nom","idx_max_nom","thin"]
heads = ["Quarter","Events","Tickets","Min (get-in)","Avg","Max","Spread max/min","Idx Min","Idx Avg","Idx Max","Thin?"]
wcol = [10,9,12,14,10,10,15,10,10,10,7]
fmt = {3:"#,##0",4:'"$"#,##0.00',5:'"$"#,##0.00',6:'"$"#,##0.00',7:"0.00",8:"0.0",9:"0.0",10:"0.0"}
write_res(wb.create_sheet("P1_Quarterly_Real"), cols_r, heads, wcol, fmt)
write_res(wb.create_sheet("P1_Quarterly_Nominal"), cols_n, heads, wcol, fmt)

wsc = wb.create_sheet("P1_Charts"); wsc["A1"] = "Indexed price lines & spread ratio"; wsc["A1"].font = TITLE
for fn, anc in [("qtr_indexed_real.png","A3"),("qtr_indexed_nominal.png","A34"),("qtr_spread_ratio.png","A65")]:
    img = XLImage(f"{DIR}/{fn}"); img.anchor = anc; wsc.add_image(img)

# ---- P2_Tier_Deltas
ws = wb.create_sheet("P2_Tier_Deltas")
ws["A1"] = "Analysis 2 — Gross share by within-year tier (2019 → 2025)"; ws["A1"].font = TITLE
dtbl = [["Tier", "2019 gross share", "2025 gross share", "Δ (pp)"]]
for lab, key in [("Top 1%","top1"),("Top 5% (cumulative)","top5"),("Mid (6–50%)","mid"),("Bottom 50%","bottom")]:
    v = P2[key]; dtbl.append([lab, v[0], v[-1], (v[-1]-v[0])])
r0 = 3
for i, row in enumerate(dtbl):
    for c, v in enumerate(row, 1):
        cell = ws.cell(r0 + i, c, v)
        if i > 0 and c in (2, 3): cell.number_format = "0.0%"
        if i > 0 and c == 4: cell.number_format = "+0.0;-0.0"
hdr(ws, 4, r0)
mid_d = (P2["mid"][-1]-P2["mid"][0])*100; top1_d = (P2["top1"][-1]-P2["top1"][0])*100
top5_d = (P2["top5"][-1]-P2["top5"][0])*100
ws.cell(r0+len(dtbl)+1, 1,
        f"Top 1% share {P2['top1'][0]*100:.1f}%→{P2['top1'][-1]*100:.1f}% ({top1_d:+.1f}pp); "
        f"Top 5% {P2['top5'][0]*100:.1f}%→{P2['top5'][-1]*100:.1f}% ({top5_d:+.1f}pp); "
        f"Mid {P2['mid'][0]*100:.1f}%→{P2['mid'][-1]*100:.1f}% ({mid_d:+.1f}pp).").font = SUB
widths(ws, [22, 16, 16, 10])

# ---- P2_Tier_ByYear (full)
ws = wb.create_sheet("P2_Tier_ByYear")
heads = ["Year","Tier","Events","Gross share","Ticket share","Avg price (real 2025$)","Avg capacity sold %"]
ws.append(heads); hdr(ws, len(heads))
for _, rr in T.iterrows():
    ws.append([rr.year, rr.tier, rr.n_events, rr.share_gross, rr.share_tickets, rr.avg_pavg_real, rr.avg_cap])
for i in range(2, len(T) + 2):
    ws.cell(i,4).number_format = "0.0%"; ws.cell(i,5).number_format = "0.0%"
    ws.cell(i,6).number_format = '"$"#,##0.00'; ws.cell(i,7).number_format = "0.0"
    if ws.cell(i,2).value == "Top 1%":
        for c in range(1,8): ws.cell(i,c).fill = PatternFill("solid", fgColor="F4CCCC")
widths(ws, [7,14,9,12,13,20,18]); ws.freeze_panes = "A2"

wsc = wb.create_sheet("P2_Tier_Chart"); wsc["A1"] = "Gross share by tier over time"; wsc["A1"].font = TITLE
img = XLImage(f"{DIR}/tier_stacked_share.png"); img.anchor = "A3"; wsc.add_image(img)

# ---- P3_Concentration
ws = wb.create_sheet("P3_Concentration")
ws["A1"] = "Analysis 3 — Revenue concentration by year"; ws["A1"].font = TITLE
heads = ["Year","Events","Headliners","Gini (events)","Gini (headliners)","HHI (headliners)","Top-10 headliner share","Total gross (USD)"]
ws.append([None]); ws.append(heads); hdr(ws, len(heads), 2)
for _, rr in C.iterrows():
    ws.append([rr.year, rr.n_events, rr.n_headliners, rr.gini_events, rr.gini_headliners,
               rr.hhi_headliners, rr.top10_head_share, rr.total_gross])
for i in range(3, len(C) + 3):
    for c in (4,5,6): ws.cell(i,c).number_format = "0.000"
    ws.cell(i,7).number_format = "0.0%"; ws.cell(i,8).number_format = '"$"#,##0'
r = len(C) + 4
ws.cell(r+1,1, f"Headliner Gini {P3['gini_h_19']:.3f} (2019) → {P3['gini_h_25']:.3f} (2025); "
              f"event Gini {P3['gini_ev_19']:.3f} → {P3['gini_ev_25']:.3f}.").font = SUB
ws.cell(r+2,1, f"Top-10 headliners' share of US gross: {P3['top10_19']*100:.1f}% (2019) → {P3['top10_25']*100:.1f}% (2025).").font = SUB
widths(ws, [7,9,11,13,15,15,20,20]); ws.freeze_panes = "A3"

wsc = wb.create_sheet("P3_Charts"); wsc["A1"] = "Concentration over time"; wsc["A1"].font = TITLE
img = XLImage(f"{DIR}/concentration.png"); img.anchor = "A3"; wsc.add_image(img)

# ---- P4_Top1_Artists
ws = wb.create_sheet("P4_Top1_Artists")
ws["A1"] = "Analysis 4 — Top 1% of artists: get-in vs top price, quarterly"; ws["A1"].font = TITLE
ws["A3"] = ("Top 1% of headliners by gross, ranked WITHIN each year. Prices attendance-weighted (ticketsSold), "
            "winsorized [p0.1,p99.9]. Ratio = weighted max ÷ weighted min (unit-free; ratio_simple uses unweighted means)."); ws["A3"].font = SUB
heads = ["Quarter","Events","Artists","Tickets","Get-in (real 2025$)","Top price (real 2025$)","Ratio (weighted)","Ratio (simple)","Thin?"]
r0 = 5; ws.append([]); ws.append([]); ws.append([]); ws.append([])
ws.cell(r0, 1)  # ensure row exists
for c, h in enumerate(heads, 1):
    ws.cell(r0, c, h)
hdr(ws, len(heads), r0)
ri = r0 + 1
for _, rr in Q4.iterrows():
    vals = [rr.quarter, rr.n_events, rr.n_artists, rr.tickets, rr.w_min_real, rr.w_max_real, rr.ratio_w, rr.ratio_s]
    for c, v in enumerate(vals, 1):
        ws.cell(ri, c, v)
    ws.cell(ri, 9, "yes" if rr.thin else "")
    ws.cell(ri, 4).number_format = "#,##0"
    ws.cell(ri, 5).number_format = '"$"#,##0.00'; ws.cell(ri, 6).number_format = '"$"#,##0.00'
    ws.cell(ri, 7).number_format = "0.00"; ws.cell(ri, 8).number_format = "0.00"
    if rr.thin:
        for c in range(1, 10): ws.cell(ri, c).fill = THIN_F
    ri += 1
ws.cell(ri + 1, 1, f"Annual weighted ratio: {P4['r19']:.2f}x (2019) → {P4['r25']:.2f}x (2025), "
                   f"{(P4['r25']/P4['r19']-1)*100:+.1f}%. Trend {P4['slope']:+.4f}x/qtr "
                   f"({'rising' if P4['slope']>0 else 'flat/declining'}).").font = SUB
ws.cell(ri + 2, 1, f"Get-in {P4['min19']:.0f}→{P4['min25']:.0f} real$; Top {P4['max19']:.0f}→{P4['max25']:.0f} real$ (2019→2025).").font = SUB
widths(ws, [10, 8, 8, 11, 18, 20, 15, 14, 8]); ws.freeze_panes = f"A{r0+1}"

wsc = wb.create_sheet("P4_Chart"); wsc["A1"] = "Top 1% artists: get-in vs top price + ratio"; wsc["A1"].font = TITLE
img = XLImage(f"{DIR}/top1_artists_ratio.png"); img.anchor = "A3"; wsc.add_image(img)

out = f"{DIR}/US_Quarterly_Ticket_Price_Spread.xlsx"
wb.save(out)
print("saved", out)
