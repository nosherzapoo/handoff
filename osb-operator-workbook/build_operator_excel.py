"""
Build a US online sports betting market-share Excel.

Operator buckets (fixed): FanDuel, DraftKings, BetMGM, Fanatics, Caesars,
ESPN Bet, Others. Barstool history is already merged into ESPN Bet via
the upstream `operator_standard` mapping.

Sheets:
  1. Handle              — $ by bucket × month, with % market-share block below
  2. GGR                 — $ by bucket × month, with % market-share block below
  3. Hold Rate           — GGR / Handle, formulas linked to sheets 1 & 2
  4. Handle Growth Y/Y   — formulas linked to sheet 1
  5. GGR Growth Y/Y      — formulas linked to sheet 2
  6. Raw Data            — flat operator-month rows with the Bucket key the
                           analytical sheets aggregate over (SUMIFS source)
  7. Coverage            — month × state matrix; each cell ✓/blank for whether
                           that state contributed online operator-level data
  8. Sources             — regulator landing pages per state
  9. Methodology         — scope, definitions, caveats

Scope: 18 operator-reporting US states, online channel only, all time.
"""

import sys
from pathlib import Path

import pandas as pd

# Standalone-handoff layout: config.py is a sibling module and the workbook is
# written next to these scripts. (In the full osb-tracker repo this module lives
# under scripts/ and imports scrapers.config; here everything is flat.)
HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

from config import STATE_REGISTRY  # noqa: E402

OPERATOR_STATES = [
    'AZ', 'CT', 'DC', 'IA', 'IL', 'IN', 'KS', 'KY', 'MA', 'MD',
    'ME', 'MI', 'MO', 'NH', 'NJ', 'NY', 'OH', 'OR', 'PA', 'WV', 'WY',
]
TARGET_BRANDS = ['FanDuel', 'DraftKings', 'BetMGM', 'Fanatics', 'Caesars', 'ESPN Bet']
# The CSV path (load_state/main) expects processed CSVs here; not shipped in the
# handoff — use build_operator_excel_from_api.py, which sources from the API.
DATA_DIR = HERE / 'data' / 'processed'
OUT_PATH = HERE / 'US_Operator_TimeSeries_Online.xlsx'


def _col_letter(idx: int) -> str:
    """1-indexed column number → Excel letter (A, B, …, Z, AA, AB, …)."""
    s = ''
    while idx:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s


def _bucket(operator_standard: str) -> str:
    """Map operator_standard to one of the 6 target brands or 'Others'."""
    if operator_standard in TARGET_BRANDS:
        return operator_standard
    return 'Others'


def load_state(st: str) -> pd.DataFrame:
    df = pd.read_csv(DATA_DIR / f'{st}.csv', low_memory=False)
    df = df[
        (df['period_type'] == 'monthly')
        & (df['channel'] == 'online')
        & (df['sport_category'].isna())
        & (df['operator_standard'].notna())
        & (~df['operator_standard'].str.upper().isin(['TOTAL', 'UNKNOWN', 'STATEWIDE', 'ALL']))
    ].copy()
    # Several scrapers (MA, MI, OH, KS, NH, plus partial OH) populate
    # gross_revenue from the regulator's published "GGR" column but never
    # backfill standard_ggr because they don't capture payouts. For
    # market-share purposes the regulator-published GGR is the canonical
    # number and matches the published hold rate, so we coalesce it into
    # standard_ggr when the latter is null.
    df['standard_ggr'] = pd.to_numeric(df['standard_ggr'], errors='coerce')
    if 'gross_revenue' in df.columns:
        gr = pd.to_numeric(df['gross_revenue'], errors='coerce')
        df['standard_ggr'] = df['standard_ggr'].fillna(gr)
    # Strict integrity rule: every row in the analytical sheets needs both
    # handle and a GGR figure. Otherwise hold-rate aggregates would mix a
    # numerator from one set of rows with a denominator from another, which
    # makes hold artificially low. Drop rows missing either.
    handle_n = pd.to_numeric(df['handle'], errors='coerce')
    df = df[handle_n.notna() & (handle_n > 0) & df['standard_ggr'].notna()].copy()
    return df[[
        'state_code', 'operator_standard', 'parent_company',
        'period_start', 'handle', 'standard_ggr', 'source_url',
    ]]


def _write_analytics_sheets(writer, brands, months):
    """Write Handle, GGR, Hold, Handle Growth Y/Y, GGR Growth Y/Y sheets.

    All sheets reference 'Raw Data' columns:
      A=State, B=Bucket, C=Parent, D=Month (date), E=Handle ($), F=GGR ($), G=Hold

    Months are written as real Excel dates so EDATE-based Y/Y formulas resolve.
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = writer.book

    title_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11, color='FFFFFF')
    section_fill = PatternFill('solid', fgColor='305496')
    others_font = Font(italic=True)
    total_font = Font(bold=True)
    total_fill = PatternFill('solid', fgColor='D9E1F2')
    header_font = Font(bold=True)
    pct_section_fill = PatternFill('solid', fgColor='548235')

    rows_per_block = len(brands) + 2  # 6 brands + Others + Total

    def _write_dollar_sheet(sheet_name, label, src_col, num_fmt='#,##0'):
        """Sheet with a dollar block at top + a % market-share block below."""
        ws = wb.create_sheet(sheet_name)
        ws.cell(row=1, column=1, value=f'{label} — US Online Sports Betting').font = title_font
        ws.cell(row=2, column=1, value=(
            'Linked dynamically to Raw Data via SUMIFS. 18 operator-reporting US states, '
            'online channel, all time. Months where a state did not publish are excluded '
            'for that state — see the Coverage sheet.'
        )).font = Font(italic=True, color='595959')

        # === DOLLAR BLOCK ===
        header_row = 4
        first_data_row = header_row + 1
        total_row = first_data_row + len(brands) + 1

        h = ws.cell(row=header_row, column=1, value=label)
        h.font = section_font; h.fill = section_fill
        for c_idx, m in enumerate(months, start=2):
            cell = ws.cell(row=header_row, column=c_idx, value=m)
            cell.font = section_font; cell.fill = section_fill
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = 'mmm yyyy'

        # 6 named brand rows
        for r_idx, brand in enumerate(brands):
            r = first_data_row + r_idx
            ws.cell(row=r, column=1, value=brand).font = header_font
            for c_idx in range(2, len(months) + 2):
                col = _col_letter(c_idx)
                f = (
                    f"=SUMIFS('Raw Data'!${src_col}:${src_col},"
                    f"'Raw Data'!$B:$B,$A{r},"
                    f"'Raw Data'!$D:$D,{col}${header_row})"
                )
                cell = ws.cell(row=r, column=c_idx, value=f)
                cell.number_format = num_fmt

        # Others row (= total minus 6 brands)
        others_row = first_data_row + len(brands)
        ws.cell(row=others_row, column=1, value='Others').font = others_font
        for c_idx in range(2, len(months) + 2):
            col = _col_letter(c_idx)
            f = (
                f"=SUMIFS('Raw Data'!${src_col}:${src_col},"
                f"'Raw Data'!$B:$B,$A{others_row},"
                f"'Raw Data'!$D:$D,{col}${header_row})"
            )
            cell = ws.cell(row=others_row, column=c_idx, value=f)
            cell.number_format = num_fmt
            cell.font = others_font

        # Total row (= sum of all 7 buckets, equivalently SUMIFS without bucket filter)
        ws.cell(row=total_row, column=1, value='Total').font = total_font
        ws.cell(row=total_row, column=1).fill = total_fill
        for c_idx in range(2, len(months) + 2):
            col = _col_letter(c_idx)
            f = f"=SUM({col}{first_data_row}:{col}{others_row})"
            cell = ws.cell(row=total_row, column=c_idx, value=f)
            cell.number_format = num_fmt
            cell.font = total_font; cell.fill = total_fill

        # === PERCENT MARKET SHARE BLOCK ===
        pct_header_row = total_row + 3
        pct_first_data_row = pct_header_row + 1

        ph = ws.cell(row=pct_header_row, column=1, value=f'{label} — % MARKET SHARE')
        ph.font = section_font; ph.fill = pct_section_fill
        for c_idx, m in enumerate(months, start=2):
            cell = ws.cell(row=pct_header_row, column=c_idx, value=m)
            cell.font = section_font; cell.fill = pct_section_fill
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = 'mmm yyyy'

        for r_idx in range(len(brands) + 1):  # 6 brands + Others
            r = pct_first_data_row + r_idx
            src_dollar_row = first_data_row + r_idx
            label_cell = ws.cell(row=r, column=1, value=ws.cell(row=src_dollar_row, column=1).value)
            label_cell.font = others_font if r_idx == len(brands) else header_font
            for c_idx in range(2, len(months) + 2):
                col = _col_letter(c_idx)
                f = f'=IFERROR({col}{src_dollar_row}/{col}{total_row},"")'
                cell = ws.cell(row=r, column=c_idx, value=f)
                cell.number_format = '0.0%'
                if r_idx == len(brands):
                    cell.font = others_font

        # Total = 100% (sanity check)
        pct_total_row = pct_first_data_row + len(brands) + 1
        tcell = ws.cell(row=pct_total_row, column=1, value='Total')
        tcell.font = total_font; tcell.fill = total_fill
        for c_idx in range(2, len(months) + 2):
            col = _col_letter(c_idx)
            f = f'=IFERROR({col}{total_row}/{col}{total_row},"")'
            cell = ws.cell(row=pct_total_row, column=c_idx, value=f)
            cell.number_format = '0.0%'
            cell.font = total_font; cell.fill = total_fill

        ws.column_dimensions['A'].width = 14
        for c_idx in range(2, len(months) + 2):
            ws.column_dimensions[_col_letter(c_idx)].width = 12
        ws.freeze_panes = 'B5'

        return header_row, first_data_row, others_row, total_row

    handle_meta = _write_dollar_sheet('Handle', 'HANDLE ($)', 'E', '#,##0')
    ggr_meta    = _write_dollar_sheet('GGR',    'GGR ($)',    'F', '#,##0')

    # === HOLD RATE SHEET (GGR / Handle, referencing the two sheets above) ===
    ws = wb.create_sheet('Hold Rate')
    ws.cell(row=1, column=1, value='HOLD RATE — US Online Sports Betting').font = title_font
    ws.cell(row=2, column=1, value='Hold = GGR / Handle. Linked to Handle and GGR sheets.').font = Font(italic=True, color='595959')

    h = ws.cell(row=4, column=1, value='HOLD %')
    h.font = section_font; h.fill = section_fill
    for c_idx, m in enumerate(months, start=2):
        cell = ws.cell(row=4, column=c_idx, value=m)
        cell.font = section_font; cell.fill = section_fill
        cell.alignment = Alignment(horizontal='center')
        cell.number_format = 'mmm yyyy'

    h_header_row, h_first_data_row, h_others_row, h_total_row = handle_meta
    g_header_row, g_first_data_row, g_others_row, g_total_row = ggr_meta

    # Same row indices line up between Handle and GGR (same shape).
    for r_idx, label in enumerate(brands + ['Others']):
        r = 5 + r_idx
        ws.cell(row=r, column=1, value=label).font = (others_font if label == 'Others' else header_font)
        src_row = h_first_data_row + r_idx  # same row in Handle and GGR
        for c_idx in range(2, len(months) + 2):
            col = _col_letter(c_idx)
            f = f"=IFERROR(GGR!{col}{src_row}/Handle!{col}{src_row},\"\")"
            cell = ws.cell(row=r, column=c_idx, value=f)
            cell.number_format = '0.0%'
            if label == 'Others':
                cell.font = others_font

    total_r = 5 + len(brands) + 1
    tcell = ws.cell(row=total_r, column=1, value='Total')
    tcell.font = total_font; tcell.fill = total_fill
    for c_idx in range(2, len(months) + 2):
        col = _col_letter(c_idx)
        f = f"=IFERROR(GGR!{col}{g_total_row}/Handle!{col}{h_total_row},\"\")"
        cell = ws.cell(row=total_r, column=c_idx, value=f)
        cell.number_format = '0.0%'
        cell.font = total_font; cell.fill = total_fill

    ws.column_dimensions['A'].width = 14
    for c_idx in range(2, len(months) + 2):
        ws.column_dimensions[_col_letter(c_idx)].width = 12
    ws.freeze_panes = 'B5'

    # === Y/Y GROWTH SHEETS ===
    def _write_growth_sheet(sheet_name, label, source_sheet, source_meta):
        ws = wb.create_sheet(sheet_name)
        ws.cell(row=1, column=1, value=f'{label} — US Online Sports Betting').font = title_font
        ws.cell(row=2, column=1, value=(
            f'Y/Y growth = current month / same month prior year − 1. Linked to {source_sheet} sheet. '
            f'Cells before the first available prior-year month are blank.'
        )).font = Font(italic=True, color='595959')

        s_header_row, s_first_data_row, s_others_row, s_total_row = source_meta

        h = ws.cell(row=4, column=1, value=label)
        h.font = section_font; h.fill = section_fill
        for c_idx, m in enumerate(months, start=2):
            cell = ws.cell(row=4, column=c_idx, value=m)
            cell.font = section_font; cell.fill = section_fill
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = 'mmm yyyy'

        for r_idx, lbl in enumerate(brands + ['Others']):
            r = 5 + r_idx
            ws.cell(row=r, column=1, value=lbl).font = (others_font if lbl == 'Others' else header_font)
            src_row = s_first_data_row + r_idx
            for c_idx in range(2, len(months) + 2):
                col = _col_letter(c_idx)
                # MATCH the date 12 months prior in the source sheet's header row.
                # Since headers are real dates, EDATE works directly.
                f = (
                    f'=IFERROR({source_sheet}!{col}{src_row}/'
                    f'INDEX({source_sheet}!{src_row}:{src_row},'
                    f'MATCH(EDATE({col}$4,-12),{source_sheet}!$4:$4,0))-1,"")'
                )
                cell = ws.cell(row=r, column=c_idx, value=f)
                cell.number_format = '0.0%'
                if lbl == 'Others':
                    cell.font = others_font

        total_r = 5 + len(brands) + 1
        tcell = ws.cell(row=total_r, column=1, value='Total')
        tcell.font = total_font; tcell.fill = total_fill
        for c_idx in range(2, len(months) + 2):
            col = _col_letter(c_idx)
            f = (
                f'=IFERROR({source_sheet}!{col}{s_total_row}/'
                f'INDEX({source_sheet}!{s_total_row}:{s_total_row},'
                f'MATCH(EDATE({col}$4,-12),{source_sheet}!$4:$4,0))-1,"")'
            )
            cell = ws.cell(row=total_r, column=c_idx, value=f)
            cell.number_format = '0.0%'
            cell.font = total_font; cell.fill = total_fill

        ws.column_dimensions['A'].width = 14
        for c_idx in range(2, len(months) + 2):
            ws.column_dimensions[_col_letter(c_idx)].width = 12
        ws.freeze_panes = 'B5'

    _write_growth_sheet('Handle Growth YoY', 'HANDLE Y/Y GROWTH', 'Handle', handle_meta)
    _write_growth_sheet('GGR Growth YoY',    'GGR Y/Y GROWTH',    'GGR',    ggr_meta)


def _write_coverage_matrix(writer, grouped, months):
    """Month × State coverage matrix: ✓ if state contributed online operator-level
    data that month, blank otherwise."""
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = writer.book
    ws = wb.create_sheet('Coverage')
    ws.cell(row=1, column=1, value='COVERAGE — Month × State').font = Font(bold=True, size=12)
    ws.cell(row=2, column=1, value=(
        'Each cell shows whether that state published online operator-level rows for that month. '
        'Months where a state is blank are excluded from that month\'s totals (denominator shrinks).'
    )).font = Font(italic=True, color='595959')

    section_fill = PatternFill('solid', fgColor='305496')
    section_font = Font(bold=True, color='FFFFFF')
    yes_fill = PatternFill('solid', fgColor='C6EFCE')

    # Header row: Month + states
    ws.cell(row=4, column=1, value='Month').font = section_font
    ws.cell(row=4, column=1).fill = section_fill
    for c_idx, st in enumerate(OPERATOR_STATES, start=2):
        cell = ws.cell(row=4, column=c_idx, value=st)
        cell.font = section_font; cell.fill = section_fill
        cell.alignment = Alignment(horizontal='center')

    # Months × states data
    state_months = grouped.groupby('state_code')['period_start'].apply(set).to_dict()

    for r_idx, m in enumerate(months, start=5):
        cell = ws.cell(row=r_idx, column=1, value=m)
        cell.number_format = 'mmm yyyy'
        cell.font = Font(bold=True)
        for c_idx, st in enumerate(OPERATOR_STATES, start=2):
            if m in state_months.get(st, set()):
                c = ws.cell(row=r_idx, column=c_idx, value='✓')
                c.alignment = Alignment(horizontal='center')
                c.fill = yes_fill

    # State count column at far right
    n_state_col = len(OPERATOR_STATES) + 2
    ws.cell(row=4, column=n_state_col, value='# States').font = section_font
    ws.cell(row=4, column=n_state_col).fill = section_fill
    for r_idx, m in enumerate(months, start=5):
        cell = ws.cell(row=r_idx, column=n_state_col, value=sum(
            1 for st in OPERATOR_STATES if m in state_months.get(st, set())
        ))
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)

    ws.column_dimensions['A'].width = 12
    for c_idx in range(2, n_state_col + 1):
        ws.column_dimensions[_col_letter(c_idx)].width = 6
    ws.freeze_panes = 'B5'


def _write_by_state_panel(writer, grouped, brands, months, sheet_name, label,
                          src_col, num_fmt='#,##0'):
    """Pivoted Operator × State panel: brand subtotal row + collapsible state
    sub-rows, month columns. Cells are SUMIFS into Raw Data where the state
    reported that month, and left EMPTY otherwise (blank = did not report, so
    the Like-for-Like Y/Y can gate on ISNUMBER). Helper columns B/C/E carry the
    machine-readable Brand/State/Level for SUMPRODUCT and are hidden.

    Returns a meta dict the Like-for-Like Y/Y writer uses to address the panel.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = writer.book
    ws = wb.create_sheet(sheet_name)

    title_font = Font(bold=True, size=12)
    section_font = Font(bold=True, color='FFFFFF')
    section_fill = PatternFill('solid', fgColor='305496')
    brand_font = Font(bold=True)
    brand_fill = PatternFill('solid', fgColor='D9E1F2')
    helper_font = Font(color='BFBFBF', size=8)

    ws.cell(row=1, column=1, value=f'{label} — by Operator × State').font = title_font
    ws.cell(row=2, column=1, value=(
        'Brand subtotal rows (bold) with collapsible state rows beneath. '
        'Blank = that state did not report that month. Linked to Raw Data via '
        'SUMIFS; feeds the Like-for-Like Y/Y sheet.'
    )).font = Font(italic=True, color='595959')

    header_row = 4
    first_data_row = 5
    month_first_col = 6

    for c_idx, title in enumerate(
            ['Operator / State', 'Brand', 'State', 'Parent', 'Lvl'], start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=title)
        cell.font = section_font
        cell.fill = section_fill
    for c_idx, m in enumerate(months, start=month_first_col):
        cell = ws.cell(row=header_row, column=c_idx, value=m)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal='center')
        cell.number_format = 'mmm yyyy'

    present = set(zip(grouped['state_code'], grouped['bucket'],
                      grouped['period_start']))
    parent_lookup = (
        grouped.dropna(subset=['parent_company'])
        .groupby(['state_code', 'bucket'])['parent_company']
        .agg(lambda s: s.mode().iat[0] if not s.mode().empty else '')
        .to_dict()
    )

    r = first_data_row
    for brand in brands:
        bstates = sorted({s for (s, b, _m) in present if b == brand})
        if not bstates:
            continue
        bmonths = {mm for (_s, b, mm) in present if b == brand}

        # Brand subtotal row (Lvl 0) — SUMIFS over all of the brand's rows.
        ws.cell(row=r, column=1, value=brand).font = brand_font
        ws.cell(row=r, column=1).fill = brand_fill
        ws.cell(row=r, column=2, value=brand).font = helper_font
        ws.cell(row=r, column=3, value='(all)').font = helper_font
        ws.cell(row=r, column=4).fill = brand_fill
        ws.cell(row=r, column=5, value=0).font = helper_font
        for c_idx, m in enumerate(months, start=month_first_col):
            cell = ws.cell(row=r, column=c_idx)
            cell.fill = brand_fill
            if m in bmonths:
                mcol = _col_letter(c_idx)
                cell.value = (
                    f"=SUMIFS('Raw Data'!${src_col}:${src_col},"
                    f"'Raw Data'!$B:$B,$B{r},'Raw Data'!$D:$D,{mcol}${header_row})"
                )
                cell.number_format = num_fmt
                cell.font = brand_font
        r += 1

        # State sub-rows (Lvl 1) — grouped/collapsible under the brand.
        for st in bstates:
            smonths = {mm for (s, b, mm) in present if b == brand and s == st}
            ws.cell(row=r, column=1, value=f'    {st}')
            ws.cell(row=r, column=2, value=brand).font = helper_font
            ws.cell(row=r, column=3, value=st).font = helper_font
            ws.cell(row=r, column=4, value=parent_lookup.get((st, brand), ''))
            ws.cell(row=r, column=5, value=1).font = helper_font
            for c_idx, m in enumerate(months, start=month_first_col):
                if m in smonths:
                    mcol = _col_letter(c_idx)
                    cell = ws.cell(row=r, column=c_idx, value=(
                        f"=SUMIFS('Raw Data'!${src_col}:${src_col},"
                        f"'Raw Data'!$A:$A,$C{r},'Raw Data'!$B:$B,$B{r},"
                        f"'Raw Data'!$D:$D,{mcol}${header_row})"
                    ))
                    cell.number_format = num_fmt
            ws.row_dimensions[r].outline_level = 1
            r += 1

    last_data_row = r - 1
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['D'].width = 16
    for col in ('B', 'C', 'E'):
        ws.column_dimensions[col].hidden = True
    for c_idx in range(month_first_col, month_first_col + len(months)):
        ws.column_dimensions[_col_letter(c_idx)].width = 12
    ws.freeze_panes = 'F5'

    return {
        'sheet': sheet_name, 'r1': first_data_row, 'r2': last_data_row,
        'hdr': header_row, 'mfc': month_first_col,
        'mlc': month_first_col + len(months) - 1,
    }


def _write_fair_yoy(writer, sheet_name, label, panel_meta, brands, months):
    """Like-for-Like Y/Y: numerator = every state reporting the current month;
    denominator = those same states a year ago (new states -> 0). Computed with
    SUMPRODUCT over the by-state panel, gated on ISNUMBER(current column)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = writer.book
    ws = wb.create_sheet(sheet_name)

    title_font = Font(bold=True, size=12)
    section_font = Font(bold=True, color='FFFFFF')
    section_fill = PatternFill('solid', fgColor='305496')
    others_font = Font(italic=True)
    total_font = Font(bold=True)
    total_fill = PatternFill('solid', fgColor='D9E1F2')
    header_font = Font(bold=True)

    P = panel_meta['sheet']
    r1, r2, hdr = panel_meta['r1'], panel_meta['r2'], panel_meta['hdr']
    mfC = _col_letter(panel_meta['mfc'])
    mlC = _col_letter(panel_meta['mlc'])

    ws.cell(row=1, column=1, value=(
        f'{label} — Like-for-Like (constant current-state panel)')).font = title_font
    ws.cell(row=2, column=1, value=(
        'Y/Y where the denominator is restricted to states that reported in the '
        'CURRENT month; newly-reporting states are still included in the '
        'numerator (numerator state-count ≥ denominator). SUMPRODUCT over the '
        f'{P} sheet. Blank until a prior-year month exists.'
    )).font = Font(italic=True, color='595959')

    header_row = 4
    ws.cell(row=header_row, column=1, value=label).font = section_font
    ws.cell(row=header_row, column=1).fill = section_fill
    for c_idx, m in enumerate(months, start=2):
        cell = ws.cell(row=header_row, column=c_idx, value=m)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal='center')
        cell.number_format = 'mmm yyyy'

    brand_rng = f"'{P}'!$B${r1}:$B${r2}"
    lvl_rng = f"'{P}'!$E${r1}:$E${r2}"

    # Each output month maps to an explicit current + prior-year panel column
    # (the panel lists the same months in the same order from column `mfc`).
    # Using fixed column letters avoids volatile INDEX/MATCH and is exact.
    month_to_panel_col = {
        (m.year, m.month): _col_letter(panel_meta['mfc'] + i)
        for i, m in enumerate(months)
    }

    def panel_cols(m):
        cur = month_to_panel_col[(m.year, m.month)]
        prior = month_to_panel_col.get((m.year - 1, m.month))
        return cur, prior

    def col_rng(letter):
        return f"'{P}'!${letter}${r1}:${letter}${r2}"

    def yoy_formula(m, brand_filter):
        cu_c, pr_c = panel_cols(m)
        if pr_c is None:
            return None
        cu, pr = col_rng(cu_c), col_rng(pr_c)
        num = f"SUMPRODUCT({brand_filter}({lvl_rng}=1)*ISNUMBER({cu})*{cu})"
        den = f"SUMPRODUCT({brand_filter}({lvl_rng}=1)*ISNUMBER({cu})*{pr})"
        return f"=IFERROR({num}/{den}-1,\"\")"

    for r_idx, brand in enumerate(brands):
        r = 5 + r_idx
        ws.cell(row=r, column=1, value=brand).font = (
            others_font if brand == 'Others' else header_font)
        for c_idx, m in enumerate(months, start=2):
            f = yoy_formula(m, f"({brand_rng}=$A{r})*")
            if f is None:
                continue
            cell = ws.cell(row=r, column=c_idx, value=f)
            cell.number_format = '0.0%'
            if brand == 'Others':
                cell.font = others_font

    total_r = 5 + len(brands)
    ws.cell(row=total_r, column=1, value='Total (US)').font = total_font
    ws.cell(row=total_r, column=1).fill = total_fill
    for c_idx, m in enumerate(months, start=2):
        f = yoy_formula(m, "")
        cell = ws.cell(row=total_r, column=c_idx)
        cell.font = total_font
        cell.fill = total_fill
        if f is None:
            continue
        cell.value = f
        cell.number_format = '0.0%'

    # Diagnostics: operator×state cells in the numerator vs. matched in the
    # denominator — makes the "numerator ≥ denominator" invariant visible.
    diag = [
        ('Reporting cells (num)',
         "SUMPRODUCT(({lvl}=1)*ISNUMBER({cu}))"),
        ('Matched prior cells (den)',
         "SUMPRODUCT(({lvl}=1)*ISNUMBER({cu})*ISNUMBER({pr}))"),
    ]
    for d_idx, (dlabel, tmpl) in enumerate(diag):
        r = total_r + 2 + d_idx
        ws.cell(row=r, column=1, value=dlabel).font = Font(italic=True, color='595959')
        for c_idx, m in enumerate(months, start=2):
            cu_c, pr_c = panel_cols(m)
            if pr_c is None:
                continue
            f = '=' + tmpl.format(lvl=lvl_rng, cu=col_rng(cu_c), pr=col_rng(pr_c))
            cell = ws.cell(row=r, column=c_idx, value=f)
            cell.number_format = '0'
            cell.font = Font(italic=True, color='595959')

    ws.column_dimensions['A'].width = 16
    for c_idx in range(2, len(months) + 2):
        ws.column_dimensions[_col_letter(c_idx)].width = 12
    ws.freeze_panes = 'B5'


def _quarter_start(ts):
    ts = pd.Timestamp(ts)
    return pd.Timestamp(ts.year, ((ts.month - 1) // 3) * 3 + 1, 1)


def _quarter_label(qs):
    return f"{(qs.month - 1) // 3 + 1}Q{qs.year % 100:02d}"


def _write_quarterly_trends(writer, months, handle_panel, ggr_panel, brands, grouped):
    """Quarterly roll-up (fully formula-driven): quarters as columns; stacked
    Handle / GGR / Handle % Y/Y / GGR % Y/Y / Hold % sections, brand rows + Total.

    Levels = SUM of the quarter's monthly cells on the Handle/GGR sheets. Y/Y is
    Like-for-Like at the quarter level: (Σ over the quarter's reported months of
    each month's constant-panel numerator) / (Σ of the same months' denominators)
    − 1, via SUMPRODUCT over the by-state panels. This is month- AND state-matched,
    so a partial trailing quarter (marked *) fairly compares only the slice
    reported so far against the identical slice a year earlier. Hold = GGR/Handle.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = writer.book
    ws = wb.create_sheet('Quarterly Trends')

    months = [pd.Timestamp(m) for m in months]
    midx = {(m.year, m.month): i for i, m in enumerate(months)}
    months_set = set(midx)

    g = grouped.copy()
    g['period_start'] = pd.to_datetime(g['period_start'])
    mcov = g.groupby('period_start')['state_code'].nunique()
    last_solid = mcov[mcov >= 0.8 * mcov.max()].index.max()

    def q_ym(qs):
        return [(qs.year, qs.month + k) for k in range(3)]

    def complete(qs):
        return all(ym in months_set and pd.Timestamp(ym[0], ym[1], 1) <= last_solid
                   for ym in q_ym(qs))

    all_q = sorted({_quarter_start(m) for m in months})
    comp = [qs for qs in all_q if complete(qs)]
    shown = list(comp)
    if comp:
        # Also show trailing (still-reporting) quarters after the last complete
        # one, marked partial — this is where "add Q2" comes in.
        last_c = max(comp)
        shown += [qs for qs in all_q
                  if qs > last_c and any(ym in months_set for ym in q_ym(qs))]
    shown = sorted(set(shown))

    # Monthly Handle/GGR sheet geometry (see _write_dollar_sheet): header row 4,
    # brand rows from 5, then Others, Total. Months from column B.
    mrow = {b: 5 + i for i, b in enumerate(brands)}
    mrow['Others'] = 5 + len(brands)
    mrow['Total'] = 5 + len(brands) + 1

    def mcol(ym):
        return _col_letter(2 + midx[ym])

    def level_formula(metric_sheet, brand, qs):
        yms = [ym for ym in q_ym(qs) if ym in months_set]
        cells = ",".join(f"'{metric_sheet}'!{mcol(ym)}{mrow[brand]}" for ym in yms)
        return f'=IF(SUM({cells})=0,"",SUM({cells}))'

    def yoy_formula(pm, brand, qs, brand_row):
        r1, r2, S = pm['r1'], pm['r2'], pm['sheet']
        lvl = f"'{S}'!$E${r1}:$E${r2}"
        bf = "" if brand == 'Total' else f"('{S}'!$B${r1}:$B${r2}=$A{brand_row})*"
        yms = [ym for ym in q_ym(qs)
               if ym in months_set and (ym[0] - 1, ym[1]) in months_set]
        if not yms:
            return None

        def rng(ym):
            return f"'{S}'!${_col_letter(pm['mfc'] + midx[ym])}${r1}:${_col_letter(pm['mfc'] + midx[ym])}${r2}"
        num = "+".join(
            f"SUMPRODUCT({bf}({lvl}=1)*ISNUMBER({rng(ym)})*{rng(ym)})" for ym in yms)
        den = "+".join(
            f"SUMPRODUCT({bf}({lvl}=1)*ISNUMBER({rng(ym)})*{rng((ym[0]-1, ym[1]))})"
            for ym in yms)
        return f'=IFERROR(({num})/({den})-1,"")'

    title_font = Font(bold=True, size=12)
    section_font = Font(bold=True, color='FFFFFF')
    section_fill = PatternFill('solid', fgColor='305496')
    total_font = Font(bold=True)
    total_fill = PatternFill('solid', fgColor='D9E1F2')
    header_font = Font(bold=True)
    others_font = Font(italic=True)

    ws.cell(row=1, column=1, value='QUARTERLY TRENDS — US Online Sports Betting').font = title_font
    ws.cell(row=2, column=1, value=(
        'Formula-driven. Y/Y is Like-for-Like: each quarter compares only the '
        'months/states reported so far against the identical slice one year '
        'earlier (Σ monthly numerators / Σ denominators). * = quarter still '
        'reporting (quarter-to-date). Handle/GGR in $.'
    )).font = Font(italic=True, color='595959')

    row_labels = list(brands) + ['Others', 'Total']

    def write_header(hr, title):
        hc = ws.cell(row=hr, column=1, value=title)
        hc.font = section_font
        hc.fill = section_fill
        for c_idx, qs in enumerate(shown, start=2):
            lbl = _quarter_label(qs) + ('*' if not complete(qs) else '')
            cc = ws.cell(row=hr, column=c_idx, value=lbl)
            cc.font = section_font
            cc.fill = section_fill
            cc.alignment = Alignment(horizontal='center')

    def style_label(r, brand):
        is_total = brand == 'Total'
        lc = ws.cell(row=r, column=1, value=brand)
        lc.font = total_font if is_total else (
            others_font if brand == 'Others' else header_font)
        if is_total:
            lc.fill = total_fill

    def apply_cell_style(cell, brand):
        if brand == 'Total':
            cell.fill = total_fill
            cell.font = total_font
        elif brand == 'Others':
            cell.font = others_font

    # Section positions (so Hold can divide GGR by Handle).
    handle_hdr = 4
    ggr_hdr = handle_hdr + len(row_labels) + 3
    hyoy_hdr = ggr_hdr + len(row_labels) + 3
    gyoy_hdr = hyoy_hdr + len(row_labels) + 3
    hold_hdr = gyoy_hdr + len(row_labels) + 3

    for hdr, title, kind in [
        (handle_hdr, 'HANDLE ($)', ('level', 'Handle')),
        (ggr_hdr, 'GGR ($)', ('level', 'GGR')),
        (hyoy_hdr, 'HANDLE % Y/Y', ('yoy', handle_panel)),
        (gyoy_hdr, 'GGR % Y/Y', ('yoy', ggr_panel)),
        (hold_hdr, 'HOLD %', ('hold', None)),
    ]:
        write_header(hdr, title)
        for r_idx, brand in enumerate(row_labels):
            r = hdr + 1 + r_idx
            style_label(r, brand)
            for c_idx, qs in enumerate(shown, start=2):
                cell = ws.cell(row=r, column=c_idx)
                apply_cell_style(cell, brand)
                if kind[0] == 'level':
                    cell.value = level_formula(kind[1], brand, qs)
                    cell.number_format = '#,##0'
                elif kind[0] == 'yoy':
                    f = yoy_formula(kind[1], brand, qs, r)
                    if f is not None:
                        cell.value = f
                        cell.number_format = '0%'
                else:  # hold = GGR / Handle for the same brand/quarter
                    col = _col_letter(c_idx)
                    gr = ggr_hdr + 1 + r_idx
                    hr_ = handle_hdr + 1 + r_idx
                    cell.value = f'=IFERROR({col}{gr}/{col}{hr_},"")'
                    cell.number_format = '0.0%'

    ws.column_dimensions['A'].width = 13
    for c_idx in range(2, len(shown) + 2):
        ws.column_dimensions[_col_letter(c_idx)].width = 14
    ws.freeze_panes = 'B5'


def build_workbook(raw: pd.DataFrame, out_path, values_in_cents: bool = True):
    """Assemble the full operator workbook from a per-operator-month frame.

    `raw` must carry columns: state_code, operator_standard, parent_company,
    period_start, handle, standard_ggr. `values_in_cents` divides handle/GGR by
    100 for the integer-cents CSV pipeline; the OSBdata API already serves
    dollars, so callers sourcing from it pass values_in_cents=False.
    """
    # Bucket mapping — anything not in the 6 named brands → Others.
    raw = raw.copy()
    raw['bucket'] = raw['operator_standard'].apply(_bucket)

    # Aggregate to (state, bucket, period_start) so the analytical sheets see one
    # row per state/bucket/month rather than per state/operator/month. Multiple
    # operators that map to the same bucket within a state-month get summed.
    parent_map = (
        raw.dropna(subset=['parent_company'])
        .groupby(['state_code', 'bucket'])['parent_company']
        .agg(lambda s: s.mode().iat[0] if not s.mode().empty else '')
        .to_dict()
    )

    grouped = (
        raw.groupby(['state_code', 'bucket', 'period_start'], as_index=False)
        .agg(handle=('handle', 'sum'), ggr=('standard_ggr', 'sum'))
    )
    # The CSV pipeline stores money as integer cents — convert to dollars for
    # the output. API-sourced callers already have dollars and skip this.
    if values_in_cents:
        grouped['handle'] = grouped['handle'] / 100.0
        grouped['ggr'] = grouped['ggr'] / 100.0
    grouped['parent_company'] = grouped.apply(
        lambda r: parent_map.get((r['state_code'], r['bucket']), ''), axis=1
    )
    grouped['hold_pct'] = grouped.apply(
        lambda r: (r['ggr'] / r['handle']) if r['handle'] and r['handle'] > 0 else None, axis=1
    )
    grouped['period_start'] = pd.to_datetime(grouped['period_start'])
    grouped = grouped.sort_values(['state_code', 'bucket', 'period_start'])

    raw_data_out = grouped.rename(columns={
        'state_code': 'State',
        'bucket': 'Bucket',
        'parent_company': 'Parent Company',
        'period_start': 'Month',
        'handle': 'Handle ($)',
        'ggr': 'GGR ($)',
        'hold_pct': 'Hold %',
    })[['State', 'Bucket', 'Parent Company', 'Month', 'Handle ($)', 'GGR ($)', 'Hold %']]

    sources_rows = []
    for st in OPERATOR_STATES:
        cfg = STATE_REGISTRY.get(st, {})
        sources_rows.append({
            'State': st,
            'State Name': cfg.get('name', ''),
            'Regulator': cfg.get('regulatory_body', ''),
            'Source Landing Page': cfg.get('source_url', ''),
            'Report Frequency': cfg.get('frequency', ''),
            'File Format': cfg.get('format', ''),
            'Launch Date': cfg.get('launch_date', ''),
        })
    sources = pd.DataFrame(sources_rows)

    notes = pd.DataFrame({
        'Note': [
            'SCOPE: Online channel only. Retail and combined channels excluded.',
            'TIME WINDOW: All time (since each state\'s online launch).',
            'STATES INCLUDED (20): AZ, CT, DC, IA, IL, IN, KS, KY, MA, MD, ME, MI, MO, NH, NY, OH, OR, PA, WV, WY.',
            'STATES EXCLUDED — no operator-level online breakdown: AR, CO, DE, LA, MS, MT, NC, NE, NV, RI, SD, TN, VA, VT.',
            'STATE EXCLUDED — structural: NJ. NJ DGE publishes operator-level GGR (from tax returns) but not operator-level handle. Without handle, hold rate cannot be computed and the row would distort the GGR-only side of any aggregate. Excluded entirely until an alternate handle source is wired in.',
            'NH and OR are DraftKings monopolies (state-exclusive contracts) — all of each state\'s online handle/GGR rolls into the DraftKings bucket.',
            'ME is a tribal-compact 2-operator market (DraftKings + Caesars).',
            'INTEGRITY RULE: every operator-month row must have BOTH a non-zero handle AND a GGR value. Rows missing either are dropped at load. This prevents hold-rate aggregates from mixing a numerator from one row-set with a denominator from another, which made hold artificially low in earlier versions of this workbook.',
            'GGR FALLBACK: when standard_ggr is null but gross_revenue is populated, gross_revenue is coalesced in. Affects MA, MI, OH, KS, NH primarily — these scrapers don\'t capture payouts so standard_ggr never derives, but the regulator-published GGR (gross_revenue) is the same number and matches the published hold rate.',
            'BUCKETS (fixed): FanDuel, DraftKings, BetMGM, Fanatics, Caesars, ESPN Bet, Others. Barstool history is already merged into ESPN Bet via the upstream operator_standard mapping.',
            'GGR DEFINITION: standard_ggr = handle − payouts. Normalized across states.',
            'HOLD: GGR / Handle, computed at the bucket level.',
            'MARKET SHARE: bucket / state-month total. States that did not publish a given month are excluded from that month\'s denominator — see Coverage sheet.',
            'Y/Y GROWTH (Reported basis): current month / same month prior year − 1, using each month AS-REPORTED. When the two months have different state coverage (e.g. a state lags this year), the growth rate is distorted. Cells before the first available prior-year month are blank.',
            'Y/Y GROWTH (Like-for-Like): the primary Y/Y. Numerator = ALL states that reported the CURRENT month; denominator = those SAME states one year earlier. A state that reported last year but is still lagging this year is dropped from BOTH sides (removes the coverage-lag distortion). A genuinely NEW state (no prior-year data) is kept in the numerator and contributes 0 to the denominator — so numerator state-count ≥ denominator state-count, and real market expansion from new states still shows as growth. Driven by SUMPRODUCT over the Handle/GGR by State panels, gated on ISNUMBER of the current-month column.',
            'BY-STATE PANELS: Handle by State / GGR by State pivot Raw Data to brand-subtotal + collapsible state rows × month. Blank = the state did not report that month (not zero). These are the SUMIFS/SUMPRODUCT source for the Like-for-Like Y/Y and support custom SUMIFS cuts.',
            'QUARTERLY TRENDS: quarters as columns; Handle / GGR / Handle % Y/Y / GGR % Y/Y / Hold % stacked by brand + Total. Fully formula-driven: levels = SUM of the quarter\'s monthly cells on the Handle/GGR sheets; Y/Y = SUMPRODUCT over the by-state panels. The trailing still-reporting quarter IS shown, marked * (quarter-to-date). Quarterly Y/Y is Like-for-Like month- AND state-matched: (Σ over the quarter\'s reported months of each month\'s current-panel numerator) / (Σ of the same months\' prior-year denominators) − 1 — so a partial quarter compares only the slice reported so far against the identical slice a year earlier.',
            'DYNAMIC: All analytical sheets use SUMIFS / INDEX-MATCH formulas pointing at Raw Data. Re-running scripts/build_operator_excel.py rewrites Raw Data; the formulas recompute on Excel open.',
            'DATA-QUALITY CAVEATS:',
            '  WV: online operators are reported as casino-skin venue names (Greenbrier, Mountaineer, Mardi Gras), not the underlying sportsbook brand. All WV rows currently fall into "Others".',
            '  IL: operator-level handle is reported statewide; operator-level GGR is sometimes sparse before mid-2023.',
            '  NY: operator-level monthly data from per-operator NYSGC PDFs.',
            '  NJ: GGR from DGE tax returns; NJ does not publish operator-level handle (only aggregate). Handle column may be sparse for NJ.',
            '  MO: online launched December 2025 — only a few months of data.',
            '  KY: 2025+ data sourced from KHRC Tableau dashboard via OCR.',
        ]
    })

    months_dt = sorted(grouped['period_start'].dt.to_pydatetime().tolist())
    months_dt = sorted(set(d.replace(day=1) for d in months_dt))

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        # Raw Data first (analytical sheets reference it).
        raw_data_out_excel = raw_data_out.copy()
        # Keep Month as a real datetime so SUMIFS by-date matching works.
        raw_data_out_excel.to_excel(writer, sheet_name='Raw Data', index=False)

        _write_analytics_sheets(writer, TARGET_BRANDS, months_dt)

        panel_brands = TARGET_BRANDS + ['Others']
        handle_panel = _write_by_state_panel(
            writer, grouped, panel_brands, months_dt,
            'Handle by State', 'HANDLE ($)', 'E')
        ggr_panel = _write_by_state_panel(
            writer, grouped, panel_brands, months_dt,
            'GGR by State', 'GGR ($)', 'F')
        _write_fair_yoy(writer, 'Handle YoY (Like-for-Like)', 'HANDLE Y/Y',
                        handle_panel, panel_brands, months_dt)
        _write_fair_yoy(writer, 'GGR YoY (Like-for-Like)', 'GGR Y/Y',
                        ggr_panel, panel_brands, months_dt)

        _write_quarterly_trends(writer, months_dt, handle_panel, ggr_panel,
                                TARGET_BRANDS, grouped)

        _write_coverage_matrix(writer, grouped, months_dt)

        sources.to_excel(writer, sheet_name='Sources', index=False)
        notes.to_excel(writer, sheet_name='Methodology', index=False)

        # Reorder: Handle / GGR / Hold / Handle Growth / GGR Growth / Raw Data / Coverage / Sources / Methodology
        wb = writer.book
        desired = [
            'Handle', 'GGR', 'Hold Rate',
            'Handle YoY (Like-for-Like)', 'GGR YoY (Like-for-Like)',
            'Quarterly Trends',
            'Handle Growth YoY', 'GGR Growth YoY',
            'Handle by State', 'GGR by State',
            'Raw Data', 'Coverage', 'Sources', 'Methodology',
        ]
        # openpyxl orders by creation; rearrange via sheet index moves.
        for target_idx, name in enumerate(desired):
            if name in wb.sheetnames:
                cur_idx = wb.sheetnames.index(name)
                offset = target_idx - cur_idx
                if offset != 0:
                    wb.move_sheet(name, offset=offset)

        # Format Raw Data — money as integers, hold as %, Month column as date
        ws = writer.sheets['Raw Data']
        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = 'yyyy-mm-dd'
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=6):
            for cell in row:
                cell.number_format = '#,##0'
        for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
            for cell in row:
                cell.number_format = '0.00%'
        # Auto-width
        for col_cells in ws.columns:
            letter = col_cells[0].column_letter
            max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
            ws.column_dimensions[letter].width = min(max_len + 2, 28)
        ws.freeze_panes = 'A2'

        # Auto-width on Sources / Methodology
        for sn in ('Sources', 'Methodology'):
            sws = writer.sheets[sn]
            for col_cells in sws.columns:
                letter = col_cells[0].column_letter
                max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
                sws.column_dimensions[letter].width = min(max_len + 2, 80)

    n_states = grouped['state_code'].nunique()
    n_buckets = grouped['bucket'].nunique()
    print(f'Wrote {out_path}')
    print(f'Raw Data rows: {len(raw_data_out)}')
    print(f'States: {n_states}, Buckets: {n_buckets}, Months: {len(months_dt)}')
    print(f'Range: {months_dt[0].strftime("%Y-%m")} to {months_dt[-1].strftime("%Y-%m")}')


def main():
    frames = [load_state(st) for st in OPERATOR_STATES]
    raw = pd.concat(frames, ignore_index=True)
    build_workbook(raw, OUT_PATH, values_in_cents=True)


if __name__ == '__main__':
    main()
